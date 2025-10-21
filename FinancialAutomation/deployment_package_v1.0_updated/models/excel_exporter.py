"""
Excel Exporter for Schedule III Financial Statements
Generates multi-sheet workbook with formula-linked financials
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any, Optional
from datetime import datetime


class ExcelExporter:
    """Export financial statements to Excel with Schedule III formatting and formula linking"""
    
    def __init__(self, company_name: str, fy_end: str):
        self.company_name = company_name
        self.fy_end = fy_end
        self.wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])
        
        # Style definitions
        self.header_font = Font(name='Bookman Old Style', size=11, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.title_font = Font(name='Bookman Old Style', size=14, bold=True)
        self.section_font = Font(name='Bookman Old Style', size=11, bold=True)
        self.normal_font = Font(name='Bookman Old Style', size=11)
        self.total_font = Font(name='Bookman Old Style', size=11, bold=True)
        self.total_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
        
        # Border styles
        thin_border = Side(style='thin', color='000000')
        self.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
        
        # Alignment
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
        self.right_align = Alignment(horizontal='right', vertical='center')
    
    def create_workbook(self, bs_data: Dict, pl_data: Dict, cf_data: Dict, notes: Dict) -> Workbook:
        """Create complete workbook with all sheets and formula linking"""
        
        # Create sheets in order
        self.create_balance_sheet(bs_data, notes)
        self.create_profit_loss(pl_data, notes)
        self.create_cash_flow(cf_data)
        self.create_notes_sheets(notes)
        
        return self.wb
    
    def create_balance_sheet(self, bs_data: Dict, notes: Dict):
        """Create Balance Sheet with formula links to Notes"""
        ws = self.wb.create_sheet('Balance Sheet', 0)
        
        # Set column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = self.company_name
        cell.font = self.title_font
        cell.alignment = self.center_align
        row += 1
        
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = f'BALANCE SHEET AS AT {self.fy_end}'
        cell.font = self.section_font
        cell.alignment = self.center_align
        row += 2
        
        # Header row
        headers = ['Particulars', 'Note', 'Current Year', 'Previous Year']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        row += 1
        
        # ASSETS
        row = self._add_bs_section(ws, row, 'ASSETS', bs_data['assets'], notes, is_asset=True)
        
        # Total Assets
        row = self._add_total_row(ws, row, 'TOTAL ASSETS', 
                                  bs_data['assets']['total_cy'], 
                                  bs_data['assets']['total_py'])
        row += 1
        
        # EQUITY AND LIABILITIES
        row = self._add_bs_section(ws, row, 'EQUITY AND LIABILITIES', 
                                   bs_data['equity_and_liabilities'], notes, is_asset=False)
        
        # Total Equity & Liabilities
        row = self._add_total_row(ws, row, 'TOTAL EQUITY AND LIABILITIES',
                                  bs_data['equity_and_liabilities']['total_cy'],
                                  bs_data['equity_and_liabilities']['total_py'])
        
        return ws
    
    def _add_bs_section(self, ws, start_row: int, section_name: str, 
                        section_data: Dict, notes: Dict, is_asset: bool) -> int:
        """Add a section of Balance Sheet with formula links"""
        row = start_row
        
        # Section header
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = section_name
        cell.font = self.section_font
        cell.alignment = self.left_align
        cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        cell.border = self.border
        row += 1
        
        # Process non-current and current sections
        for subsection_key in ['non_current', 'current'] if is_asset else ['equity', 'non_current_liabilities', 'current_liabilities']:
            if subsection_key not in section_data:
                continue
            
            subsection = section_data[subsection_key]
            subsection_name = self._format_subsection_name(subsection_key)
            
            # Subsection header
            cell = ws[f'A{row}']
            cell.value = subsection_name
            cell.font = self.section_font
            cell.alignment = self.left_align
            row += 1
            
            # Line items
            for key, value in subsection.items():
                if key in ['total_cy', 'total_py']:
                    continue
                
                if isinstance(value, dict) and 'cy' in value and 'py' in value:
                    # Find note number for this line item
                    note_ref = self._find_note_reference(key, notes)
                    
                    # Item name
                    item_name = self._format_line_item_name(key)
                    ws[f'A{row}'] = item_name
                    ws[f'A{row}'].font = self.normal_font
                    ws[f'A{row}'].alignment = self.left_align
                    ws[f'A{row}'].border = self.border
                    
                    # Note reference
                    ws[f'B{row}'] = note_ref if note_ref else ''
                    ws[f'B{row}'].font = self.normal_font
                    ws[f'B{row}'].alignment = self.center_align
                    ws[f'B{row}'].border = self.border
                    
                    # Current Year - Formula link to Notes sheet
                    if note_ref:
                        formula = f"='Note_{note_ref}'!C{self._get_note_total_row(note_ref)}"
                        ws[f'C{row}'] = formula
                    else:
                        ws[f'C{row}'] = value['cy']
                    ws[f'C{row}'].font = self.normal_font
                    ws[f'C{row}'].alignment = self.right_align
                    ws[f'C{row}'].number_format = '#,##0.00'
                    ws[f'C{row}'].border = self.border
                    
                    # Previous Year
                    if note_ref:
                        formula = f"='Note_{note_ref}'!D{self._get_note_total_row(note_ref)}"
                        ws[f'D{row}'] = formula
                    else:
                        ws[f'D{row}'] = value['py']
                    ws[f'D{row}'].font = self.normal_font
                    ws[f'D{row}'].alignment = self.right_align
                    ws[f'D{row}'].number_format = '#,##0.00'
                    ws[f'D{row}'].border = self.border
                    
                    row += 1
            
            # Subsection total
            if 'total_cy' in subsection and 'total_py' in subsection:
                row = self._add_subtotal_row(ws, row, f'Total {subsection_name}',
                                            subsection['total_cy'],
                                            subsection['total_py'])
        
        return row
    
    def _add_total_row(self, ws, row: int, label: str, cy_value: float, py_value: float) -> int:
        """Add a total row with bold formatting"""
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = self.total_font
        ws[f'A{row}'].alignment = self.left_align
        ws[f'A{row}'].fill = self.total_fill
        ws[f'A{row}'].border = self.border
        
        ws[f'B{row}'] = ''
        ws[f'B{row}'].border = self.border
        ws[f'B{row}'].fill = self.total_fill
        
        ws[f'C{row}'] = cy_value
        ws[f'C{row}'].font = self.total_font
        ws[f'C{row}'].alignment = self.right_align
        ws[f'C{row}'].number_format = '#,##0.00'
        ws[f'C{row}'].fill = self.total_fill
        ws[f'C{row}'].border = self.border
        
        ws[f'D{row}'] = py_value
        ws[f'D{row}'].font = self.total_font
        ws[f'D{row}'].alignment = self.right_align
        ws[f'D{row}'].number_format = '#,##0.00'
        ws[f'D{row}'].fill = self.total_fill
        ws[f'D{row}'].border = self.border
        
        return row + 1
    
    def _add_subtotal_row(self, ws, row: int, label: str, cy_value: float, py_value: float) -> int:
        """Add a subtotal row"""
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = self.section_font
        ws[f'A{row}'].alignment = self.left_align
        ws[f'A{row}'].border = self.border
        
        ws[f'B{row}'] = ''
        ws[f'B{row}'].border = self.border
        
        ws[f'C{row}'] = cy_value
        ws[f'C{row}'].font = self.section_font
        ws[f'C{row}'].alignment = self.right_align
        ws[f'C{row}'].number_format = '#,##0.00'
        ws[f'C{row}'].border = self.border
        
        ws[f'D{row}'] = py_value
        ws[f'D{row}'].font = self.section_font
        ws[f'D{row}'].alignment = self.right_align
        ws[f'D{row}'].number_format = '#,##0.00'
        ws[f'D{row}'].border = self.border
        
        return row + 1
    
    def create_profit_loss(self, pl_data: Dict, notes: Dict):
        """Create Profit & Loss Statement with formula links"""
        ws = self.wb.create_sheet('Profit & Loss')
        
        # Set column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = self.company_name
        cell.font = self.title_font
        cell.alignment = self.center_align
        row += 1
        
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = f'STATEMENT OF PROFIT AND LOSS FOR THE YEAR ENDED {self.fy_end}'
        cell.font = self.section_font
        cell.alignment = self.center_align
        row += 2
        
        # Header row
        headers = ['Particulars', 'Note', 'Current Year', 'Previous Year']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        row += 1
        
        # Revenue section
        ws[f'A{row}'] = 'I. Revenue from Operations'
        ws[f'A{row}'].font = self.section_font
        row += 1
        
        ws[f'A{row}'] = 'Revenue from Operations'
        ws[f'C{row}'] = pl_data['revenue']['cy']
        ws[f'D{row}'] = pl_data['revenue']['py']
        self._apply_row_formatting(ws, row)
        row += 1
        
        ws[f'A{row}'] = 'II. Other Income'
        ws[f'C{row}'] = pl_data['other_income']['cy']
        ws[f'D{row}'] = pl_data['other_income']['py']
        self._apply_row_formatting(ws, row)
        row += 1
        
        ws[f'A{row}'] = 'III. Total Income (I + II)'
        ws[f'C{row}'] = pl_data['total_income_cy']
        ws[f'D{row}'] = pl_data['total_income_py']
        self._apply_row_formatting(ws, row, bold=True)
        row += 2
        
        # Expenses section
        ws[f'A{row}'] = 'IV. EXPENSES'
        ws[f'A{row}'].font = self.section_font
        row += 1
        
        for key, value in pl_data['expenses'].items():
            if isinstance(value, dict) and 'cy' in value:
                ws[f'A{row}'] = self._format_line_item_name(key)
                ws[f'C{row}'] = value['cy']
                ws[f'D{row}'] = value['py']
                self._apply_row_formatting(ws, row)
                row += 1
        
        ws[f'A{row}'] = 'Total Expenses (IV)'
        ws[f'C{row}'] = pl_data['total_expenses_cy']
        ws[f'D{row}'] = pl_data['total_expenses_py']
        self._apply_row_formatting(ws, row, bold=True)
        row += 2
        
        # Profit calculations
        ws[f'A{row}'] = 'V. Profit/(Loss) before Tax (III - IV)'
        ws[f'C{row}'] = pl_data['profit_before_tax_cy']
        ws[f'D{row}'] = pl_data['profit_before_tax_py']
        self._apply_row_formatting(ws, row, bold=True)
        row += 2
        
        ws[f'A{row}'] = 'VI. Tax Expense'
        ws[f'C{row}'] = pl_data['tax_cy']
        ws[f'D{row}'] = pl_data['tax_py']
        self._apply_row_formatting(ws, row)
        row += 2
        
        ws[f'A{row}'] = 'VII. Profit/(Loss) for the Year (V - VI)'
        ws[f'C{row}'] = pl_data['profit_after_tax_cy']
        ws[f'D{row}'] = pl_data['profit_after_tax_py']
        self._apply_row_formatting(ws, row, bold=True, highlight=True)
        
        return ws
    
    def create_cash_flow(self, cf_data: Dict):
        """Create Cash Flow Statement"""
        ws = self.wb.create_sheet('Cash Flow')
        
        # Set column widths
        ws.column_dimensions['A'].width = 55
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = self.company_name
        cell.font = self.title_font
        cell.alignment = self.center_align
        row += 1
        
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = f'CASH FLOW STATEMENT FOR THE YEAR ENDED {self.fy_end}'
        cell.font = self.section_font
        cell.alignment = self.center_align
        row += 1
        
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = '(Indirect Method)'
        cell.font = self.normal_font
        cell.alignment = self.center_align
        row += 2
        
        # Header row
        headers = ['Particulars', 'Current Year', 'Previous Year']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        row += 1
        
        # Operating Activities
        row = self._add_cash_flow_section(ws, row, 'A. CASH FLOW FROM OPERATING ACTIVITIES',
                                          cf_data['operating_activities'])
        
        # Investing Activities
        row = self._add_cash_flow_section(ws, row, 'B. CASH FLOW FROM INVESTING ACTIVITIES',
                                          cf_data['investing_activities'])
        
        # Financing Activities
        row = self._add_cash_flow_section(ws, row, 'C. CASH FLOW FROM FINANCING ACTIVITIES',
                                          cf_data['financing_activities'])
        
        # Net increase/decrease
        row += 1
        ws[f'A{row}'] = 'Net Increase/(Decrease) in Cash and Cash Equivalents (A+B+C)'
        ws[f'B{row}'] = cf_data['net_increase_cy']
        ws[f'C{row}'] = cf_data['net_increase_py']
        self._apply_row_formatting(ws, row, bold=True, highlight=True)
        row += 1
        
        ws[f'A{row}'] = 'Cash and Cash Equivalents at Beginning of Year'
        ws[f'B{row}'] = cf_data['opening_cash_cy']
        ws[f'C{row}'] = cf_data['opening_cash_py']
        self._apply_row_formatting(ws, row)
        row += 1
        
        ws[f'A{row}'] = 'Cash and Cash Equivalents at End of Year'
        ws[f'B{row}'] = cf_data['closing_cash_cy']
        ws[f'C{row}'] = cf_data['closing_cash_py']
        self._apply_row_formatting(ws, row, bold=True, highlight=True)
        
        return ws
    
    def _add_cash_flow_section(self, ws, start_row: int, section_name: str, section_data: Dict) -> int:
        """Add a Cash Flow section"""
        row = start_row
        
        # Section header
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = section_name
        cell.font = self.section_font
        cell.alignment = self.left_align
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        cell.border = self.border
        row += 1
        
        # Extract net cash (the summary value)
        net_cash_cy = section_data.get('net_cash_cy', 0)
        net_cash_py = section_data.get('net_cash_py', 0)
        
        # Add simple line item for now (detailed breakdown can be added later)
        ws[f'A{row}'] = f'  Net Cash from {section_name.split("FROM")[-1].strip()}'
        ws[f'B{row}'] = net_cash_cy
        ws[f'C{row}'] = net_cash_py
        self._apply_row_formatting(ws, row, bold=True)
        row += 1
        
        return row
    
    def create_notes_sheets(self, notes: Dict):
        """Create individual sheets for each note"""
        for note_num, note_data in sorted(notes.items()):
            self._create_note_sheet(note_num, note_data)
    
    def _create_note_sheet(self, note_num: int, note_data: Dict):
        """Create a single note sheet"""
        sheet_name = f'Note_{note_num}'
        ws = self.wb.create_sheet(sheet_name)
        
        # Set column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 5
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = note_data['title']
        cell.font = self.title_font
        cell.alignment = self.center_align
        row += 2
        
        # Header row
        headers = ['Particulars', '', 'Current Year', 'Previous Year']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        row += 1
        
        # Data rows
        if 'data' in note_data and note_data['data']:
            row = self._add_note_data_rows(ws, row, note_data['data'])
        
        # Total row
        ws[f'A{row}'] = 'TOTAL'
        ws[f'A{row}'].font = self.total_font
        ws[f'A{row}'].alignment = self.left_align
        ws[f'A{row}'].fill = self.total_fill
        ws[f'A{row}'].border = self.border
        
        ws[f'B{row}'].border = self.border
        ws[f'B{row}'].fill = self.total_fill
        
        ws[f'C{row}'] = note_data.get('total_cy', 0)
        ws[f'C{row}'].font = self.total_font
        ws[f'C{row}'].alignment = self.right_align
        ws[f'C{row}'].number_format = '#,##0.00'
        ws[f'C{row}'].fill = self.total_fill
        ws[f'C{row}'].border = self.border
        
        ws[f'D{row}'] = note_data.get('total_py', 0)
        ws[f'D{row}'].font = self.total_font
        ws[f'D{row}'].alignment = self.right_align
        ws[f'D{row}'].number_format = '#,##0.00'
        ws[f'D{row}'].fill = self.total_fill
        ws[f'D{row}'].border = self.border
        
        # Store total row for formula references
        self._note_total_rows[note_num] = row
        
        return ws
    
    def _add_note_data_rows(self, ws, start_row: int, data: Dict, indent: int = 0) -> int:
        """Add data rows for a note (handles nested structures)"""
        row = start_row
        
        for key, value in data.items():
            if isinstance(value, dict):
                if 'cy' in value and 'py' in value:
                    # Simple data row
                    ws[f'A{row}'] = '  ' * indent + self._format_line_item_name(key)
                    ws[f'A{row}'].font = self.normal_font
                    ws[f'A{row}'].alignment = self.left_align
                    ws[f'A{row}'].border = self.border
                    
                    ws[f'B{row}'].border = self.border
                    
                    ws[f'C{row}'] = value['cy']
                    ws[f'C{row}'].font = self.normal_font
                    ws[f'C{row}'].alignment = self.right_align
                    ws[f'C{row}'].number_format = '#,##0.00'
                    ws[f'C{row}'].border = self.border
                    
                    ws[f'D{row}'] = value['py']
                    ws[f'D{row}'].font = self.normal_font
                    ws[f'D{row}'].alignment = self.right_align
                    ws[f'D{row}'].number_format = '#,##0.00'
                    ws[f'D{row}'].border = self.border
                    
                    row += 1
                else:
                    # Nested section (e.g., ageing breakdown)
                    ws.merge_cells(f'A{row}:D{row}')
                    cell = ws[f'A{row}']
                    cell.value = '  ' * indent + self._format_line_item_name(key)
                    cell.font = self.section_font
                    cell.alignment = self.left_align
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                    cell.border = self.border
                    row += 1
                    
                    # Recursively add nested data
                    row = self._add_note_data_rows(ws, row, value, indent + 1)
        
        return row
    
    def _apply_row_formatting(self, ws, row: int, bold: bool = False, highlight: bool = False):
        """Apply standard row formatting"""
        font = self.total_font if bold else self.normal_font
        fill = self.total_fill if highlight else None
        
        for col in ['A', 'B', 'C', 'D']:
            if col in ['C', 'D']:
                ws[f'{col}{row}'].font = font
                ws[f'{col}{row}'].alignment = self.right_align
                ws[f'{col}{row}'].number_format = '#,##0.00'
            else:
                ws[f'{col}{row}'].font = font
                ws[f'{col}{row}'].alignment = self.left_align
            
            ws[f'{col}{row}'].border = self.border
            if fill:
                ws[f'{col}{row}'].fill = fill
    
    def _format_subsection_name(self, key: str) -> str:
        """Format subsection name for display"""
        mapping = {
            'non_current': 'Non-Current Assets',
            'current': 'Current Assets',
            'equity': 'Equity',
            'non_current_liabilities': 'Non-Current Liabilities',
            'current_liabilities': 'Current Liabilities'
        }
        return mapping.get(key, key.replace('_', ' ').title())
    
    def _format_line_item_name(self, key: str) -> str:
        """Format line item name for display"""
        # Special mappings
        special_names = {
            'ppe': 'Property, Plant and Equipment',
            'cwip': 'Capital Work-in-Progress',
            'cash_and_bank': 'Cash and Cash Equivalents',
            'trade_receivables': 'Trade Receivables',
            'trade_payables': 'Trade Payables',
            'inventories': 'Inventories',
            'share_capital': 'Share Capital',
            'other_equity': 'Other Equity',
            'borrowings': 'Borrowings',
            'deferred_tax': 'Deferred Tax',
            'provisions': 'Provisions',
            'secured': 'Secured',
            'unsecured_good': 'Unsecured - Considered Good',
            'unsecured_doubtful': 'Unsecured - Credit Impaired',
            'allowance_doubtful': 'Less: Allowance for Doubtful Debts',
            'outstanding_0_6months': '0-6 months',
            'outstanding_6_12months': '6-12 months',
            'outstanding_1_2years': '1-2 years',
            'outstanding_2_3years': '2-3 years',
            'outstanding_gt_3years': 'More than 3 years',
            'msme': 'Micro and Small Enterprises',
            'others': 'Other than MSME',
            'cash_on_hand': 'Cash on Hand',
            'balances_with_banks': 'Balances with Banks',
            'cheques_on_hand': 'Cheques on Hand'
        }
        
        if key in special_names:
            return special_names[key]
        
        # Default: Replace underscores and title case
        return key.replace('_', ' ').title()
    
    def _find_note_reference(self, line_item_key: str, notes: Dict) -> Optional[int]:
        """Find the note number that corresponds to a Balance Sheet line item"""
        # Mapping of BS line items to note numbers
        mappings = {
            'ppe': 1,
            'cwip': 2,
            'non_current_investments': 3,
            'loans': 4,
            'other_financial_assets': 5,
            'deferred_tax_assets': 6,
            'other_non_current_assets': 7,
            'inventories': 8,
            'current_investments': 9,
            'trade_receivables': 10,
            'cash_and_bank': 11,
            'bank_balances': 12,
            'current_loans': 13,
            'other_current_financial_assets': 14,
            'other_current_assets': 15,
            'share_capital': 16,
            'other_equity': 17,
            'long_term_borrowings': 18,
            'other_non_current_liabilities': 19,
            'non_current_provisions': 20,
            'deferred_tax_liabilities': 21,
            'other_nc_liabilities': 22,
            'short_term_borrowings': 23,
            'trade_payables': 24,
            'other_current_liabilities': 26,
            'current_provisions': 27
        }
        
        return mappings.get(line_item_key)
    
    def _get_note_total_row(self, note_num: int) -> int:
        """Get the row number where the total appears in a note sheet"""
        # Will be populated when creating note sheets
        if not hasattr(self, '_note_total_rows'):
            self._note_total_rows = {}
        return self._note_total_rows.get(note_num, 10)  # Default to row 10
    
    def save(self, filename: str):
        """Save the workbook to a file"""
        self.wb.save(filename)
        return filename
