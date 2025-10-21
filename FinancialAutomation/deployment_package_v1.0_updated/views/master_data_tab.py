"""Master Data Management Tab - Complete CRUD Interface"""

from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton, 
                             QLabel, QLineEdit, QComboBox, QTreeWidget, 
                             QTreeWidgetItem, QMessageBox, QDialog, QFormLayout,
                             QGroupBox, QSplitter, QTextEdit, QFileDialog)
from PyQt5.QtCore import Qt
from models.master_data import MajorHead, MinorHead, Grouping
import openpyxl
from openpyxl.styles import Font, PatternFill
import sqlite3
import os

class MasterDataTab(QWidget):
    """Master Data Management Tab with Full CRUD Operations"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_window = parent
        self.current_level = "major"  # major, minor, or grouping
        self.selected_major_id = None
        self.selected_minor_id = None
        self.selected_grouping_id = None
        self.selected_item_id = None  # Currently selected item for update/delete
        self.init_ui()
        self.load_data()
    
    @property
    def current_company_id(self):
        """Get current company ID from parent window"""
        if hasattr(self.parent_window, 'current_company_id'):
            return self.parent_window.current_company_id
        return None
    
    def init_ui(self):
        """Initialize UI with tree view and CRUD controls"""
        main_layout = QHBoxLayout()
        
        # Create splitter for resizable sections
        splitter = QSplitter(Qt.Horizontal)
        
        # Left panel - Tree view
        left_panel = self.create_tree_panel()
        splitter.addWidget(left_panel)
        
        # Right panel - CRUD operations
        right_panel = self.create_crud_panel()
        splitter.addWidget(right_panel)
        
        # Set initial sizes (60% tree, 40% CRUD)
        splitter.setSizes([600, 400])
        
        main_layout.addWidget(splitter)
        self.setLayout(main_layout)
    
    def create_tree_panel(self):
        """Create tree view panel"""
        panel = QWidget()
        layout = QVBoxLayout()
        
        # Header
        header_layout = QHBoxLayout()
        title = QLabel("Chart of Accounts - Hierarchical View")
        title.setStyleSheet("font-weight: bold; font-size: 14pt;")
        header_layout.addWidget(title)
        header_layout.addStretch()
        
        # Refresh button
        refresh_btn = QPushButton("🔄 Refresh")
        refresh_btn.clicked.connect(self.load_data)
        header_layout.addWidget(refresh_btn)
        
        layout.addLayout(header_layout)
        
        # Tree widget
        self.tree = QTreeWidget()
        self.tree.setHeaderLabels(["Name", "Code", "Type", "ID"])
        self.tree.setColumnWidth(0, 300)
        self.tree.setColumnWidth(1, 100)
        self.tree.setColumnWidth(2, 100)
        self.tree.itemClicked.connect(self.on_tree_item_clicked)
        layout.addWidget(self.tree)
        
        # Statistics
        self.stats_label = QLabel()
        self.stats_label.setStyleSheet("padding: 5px; background-color: #f0f0f0;")
        layout.addWidget(self.stats_label)
        
        panel.setLayout(layout)
        return panel
    
    def create_crud_panel(self):
        """Create CRUD operations panel"""
        panel = QWidget()
        layout = QVBoxLayout()
        
        # Level selector
        level_group = QGroupBox("Select Level")
        level_layout = QVBoxLayout()
        
        self.level_combo = QComboBox()
        self.level_combo.addItems(["Major Head", "Minor Head", "Grouping"])
        self.level_combo.currentTextChanged.connect(self.on_level_changed)
        level_layout.addWidget(self.level_combo)
        
        level_group.setLayout(level_layout)
        layout.addWidget(level_group)
        
        # Form group
        form_group = QGroupBox("Details")
        form_layout = QFormLayout()
        
        # Name field
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("Enter name...")
        form_layout.addRow("Name*:", self.name_input)
        
        # Code field
        self.code_input = QLineEdit()
        self.code_input.setPlaceholderText("Enter code...")
        form_layout.addRow("Code*:", self.code_input)
        
        # Parent selectors (conditional)
        self.parent_major_combo = QComboBox()
        self.parent_major_label = QLabel("Major Head*:")
        form_layout.addRow(self.parent_major_label, self.parent_major_combo)
        
        self.parent_minor_combo = QComboBox()
        self.parent_minor_label = QLabel("Minor Head*:")
        form_layout.addRow(self.parent_minor_label, self.parent_minor_combo)
        
        # Description
        self.description_input = QTextEdit()
        self.description_input.setPlaceholderText("Enter description (optional)...")
        self.description_input.setMaximumHeight(80)
        form_layout.addRow("Description:", self.description_input)
        
        # Opening Balance CY
        from PyQt5.QtWidgets import QDoubleSpinBox
        self.balance_cy_input = QDoubleSpinBox()
        self.balance_cy_input.setRange(-999999999999.99, 999999999999.99)
        self.balance_cy_input.setDecimals(2)
        self.balance_cy_input.setValue(0.0)
        self.balance_cy_input.setPrefix("₹ ")
        form_layout.addRow("Opening Balance (CY):", self.balance_cy_input)
        
        # Opening Balance PY
        self.balance_py_input = QDoubleSpinBox()
        self.balance_py_input.setRange(-999999999999.99, 999999999999.99)
        self.balance_py_input.setDecimals(2)
        self.balance_py_input.setValue(0.0)
        self.balance_py_input.setPrefix("₹ ")
        form_layout.addRow("Opening Balance (PY):", self.balance_py_input)
        
        form_group.setLayout(form_layout)
        layout.addWidget(form_group)
        
        # Action buttons
        btn_layout = QHBoxLayout()
        
        self.add_btn = QPushButton("➕ Add New")
        self.add_btn.clicked.connect(self.add_item)
        self.add_btn.setStyleSheet("background-color: #4CAF50; color: white; padding: 8px;")
        btn_layout.addWidget(self.add_btn)
        
        self.update_btn = QPushButton("✏️ Update")
        self.update_btn.clicked.connect(self.update_item)
        self.update_btn.setEnabled(False)
        self.update_btn.setStyleSheet("background-color: #2196F3; color: white; padding: 8px;")
        btn_layout.addWidget(self.update_btn)
        
        self.delete_btn = QPushButton("🗑️ Delete")
        self.delete_btn.clicked.connect(self.delete_item)
        self.delete_btn.setEnabled(False)
        self.delete_btn.setStyleSheet("background-color: #f44336; color: white; padding: 8px;")
        btn_layout.addWidget(self.delete_btn)
        
        self.clear_btn = QPushButton("🔄 Clear")
        self.clear_btn.clicked.connect(self.clear_form)
        btn_layout.addWidget(self.clear_btn)
        
        layout.addLayout(btn_layout)
        
        # Import/Export section
        ie_group = QGroupBox("Import / Export")
        ie_layout = QHBoxLayout()
        
        import_btn = QPushButton("📥 Import Excel")
        import_btn.clicked.connect(self.import_from_excel)
        ie_layout.addWidget(import_btn)
        
        export_btn = QPushButton("📤 Export Excel")
        export_btn.clicked.connect(self.export_to_excel)
        ie_layout.addWidget(export_btn)
        
        ie_group.setLayout(ie_layout)
        layout.addWidget(ie_group)
        
        # Help text
        help_text = QLabel(
            "💡 <b>Tips:</b><br>"
            "• Click tree items to edit/delete<br>"
            "• Major → Minor → Grouping hierarchy<br>"
            "• Import/Export for bulk operations<br>"
            "• All fields marked * are required"
        )
        help_text.setWordWrap(True)
        help_text.setStyleSheet("padding: 10px; background-color: #fff3cd; border-radius: 5px;")
        layout.addWidget(help_text)
        
        layout.addStretch()
        
        panel.setLayout(layout)
        return panel
    
    def on_level_changed(self, level_text):
        """Handle level selection change"""
        level_map = {
            "Major Head": "major",
            "Minor Head": "minor",
            "Grouping": "grouping"
        }
        self.current_level = level_map[level_text]
        self.update_form_visibility()
        self.clear_form()
        self.load_parent_combos()
    
    def update_form_visibility(self):
        """Update form field visibility based on selected level"""
        if self.current_level == "major":
            self.parent_major_combo.hide()
            self.parent_major_label.hide()
            self.parent_minor_combo.hide()
            self.parent_minor_label.hide()
        elif self.current_level == "minor":
            self.parent_major_combo.show()
            self.parent_major_label.show()
            self.parent_minor_combo.hide()
            self.parent_minor_label.hide()
        else:  # grouping
            self.parent_major_combo.show()
            self.parent_major_label.show()
            self.parent_minor_combo.show()
            self.parent_minor_label.show()
    
    def load_parent_combos(self):
        """Load parent combo boxes"""
        if not self.current_company_id:
            return
        
        # Load major heads for current company
        self.parent_major_combo.clear()
        major_heads = MajorHead.get_all_by_company(self.current_company_id)
        for major in major_heads:
            self.parent_major_combo.addItem(f"{major.major_head_name}", major.major_head_id)
        
        # Connect signal to update minor heads when major changes
        try:
            self.parent_major_combo.currentIndexChanged.disconnect()
        except:
            pass
        self.parent_major_combo.currentIndexChanged.connect(self.load_minor_heads_combo)
        
        # Initial load of minor heads
        self.load_minor_heads_combo()
    
    def load_minor_heads_combo(self):
        """Load minor heads based on selected major head"""
        self.parent_minor_combo.clear()
        
        if not self.current_company_id:
            return
        
        if self.current_level == "grouping":
            major_id = self.parent_major_combo.currentData()
            if major_id:
                minor_heads = MinorHead.get_all(company_id=self.current_company_id, major_head_id=major_id)
                for minor in minor_heads:
                    # minor tuple: (minor_head_id, company_id, major_head_id, minor_head_name, ...)
                    self.parent_minor_combo.addItem(f"{minor[3]}", minor[0])
    
    def load_data(self):
        """Load all data into tree view"""
        self.tree.clear()
        
        if not self.current_company_id:
            self.stats_label.setText("⚠️ No company selected. Please select a company first.")
            self.stats_label.setStyleSheet("padding: 5px; background-color: #ffe0e0; color: #d32f2f;")
            return
        
        major_count = 0
        minor_count = 0
        grouping_count = 0
        
        # Load major heads for current company
        major_heads = MajorHead.get_all_by_company(self.current_company_id)
        for major in major_heads:
            major_count += 1
            
            major_item = QTreeWidgetItem(self.tree)
            major_item.setText(0, major.major_head_name)
            major_item.setText(1, major.category or "")
            major_item.setText(2, "Major")
            major_item.setText(3, str(major.major_head_id))
            major_item.setForeground(0, Qt.blue)
            major_item.setData(0, Qt.UserRole, {"type": "major", "id": major.major_head_id})
            
            # Load minor heads for this major
            minor_heads = MinorHead.get_all(company_id=self.current_company_id, major_head_id=major.major_head_id)
            for minor in minor_heads:
                # minor tuple: (minor_head_id, company_id, major_head_id, minor_head_name, opening_balance_cy, opening_balance_py, code, description)
                minor_id, _, _, minor_name, _, _, minor_code, _ = minor
                minor_count += 1
                
                minor_item = QTreeWidgetItem(major_item)
                minor_item.setText(0, minor_name)
                minor_item.setText(1, minor_code)
                minor_item.setText(2, "Minor")
                minor_item.setText(3, str(minor_id))
                minor_item.setForeground(0, Qt.darkGreen)
                minor_item.setData(0, Qt.UserRole, {"type": "minor", "id": minor_id, "major_id": major.major_head_id})
                
                # Load groupings for this minor
                groupings = Grouping.get_all(company_id=self.current_company_id, minor_head_id=minor_id)
                for grouping in groupings:
                    # grouping tuple: (grouping_id, company_id, minor_head_id, grouping_name, opening_balance_cy, opening_balance_py, code, description)
                    grouping_id, _, _, group_name, _, _, group_code, _ = grouping
                    grouping_count += 1
                    
                    grouping_item = QTreeWidgetItem(minor_item)
                    grouping_item.setText(0, group_name)
                    grouping_item.setText(1, group_code)
                    grouping_item.setText(2, "Grouping")
                    grouping_item.setText(3, str(grouping_id))
                    grouping_item.setForeground(0, Qt.darkMagenta)
                    grouping_item.setData(0, Qt.UserRole, {"type": "grouping", "id": grouping_id, "minor_id": minor_id})
        
        # Expand all
        self.tree.expandAll()
        
        # Update statistics
        self.stats_label.setText(
            f"📊 Total: {major_count} Major Heads | {minor_count} Minor Heads | {grouping_count} Groupings"
        )
        self.stats_label.setStyleSheet("padding: 5px; background-color: #e8f5e9; color: #2e7d32;")
        
        # Reload parent combos
        self.load_parent_combos()
    
    def on_tree_item_clicked(self, item, column):
        """Handle tree item click"""
        data = item.data(0, Qt.UserRole)
        if not data:
            return
        
        item_type = data["type"]
        item_id = data["id"]
        self.selected_item_id = item_id
        
        # Set level combo
        level_map = {"major": 0, "minor": 1, "grouping": 2}
        self.level_combo.setCurrentIndex(level_map[item_type])
        
        # Load item data
        if item_type == "major":
            major = MajorHead.get_by_id(item_id)
            if major:
                # major tuple: (major_head_id, major_head_name, category, description)
                self.name_input.setText(major[1])
                self.code_input.setText(major[2] or "")
                self.description_input.setText(major[3] or "")
                # Get full major object for balances
                major_obj_list = [m for m in MajorHead.get_all_by_company(self.current_company_id) if m.major_head_id == item_id]
                if major_obj_list:
                    major_obj = major_obj_list[0]
                    self.balance_cy_input.setValue(major_obj.opening_balance_cy or 0.0)
                    self.balance_py_input.setValue(major_obj.opening_balance_py or 0.0)
                self.selected_major_id = item_id
        
        elif item_type == "minor":
            minor = MinorHead.get_by_id(item_id)
            if minor:
                # minor tuple: (minor_head_id, major_head_id, minor_head_name, code, description)
                self.name_input.setText(minor[2])
                self.code_input.setText(minor[3] or "")
                self.description_input.setText(minor[4] or "")
                # Get full minor for balances
                minor_full = MinorHead.get_all(company_id=self.current_company_id, major_head_id=None)
                for m in minor_full:
                    if m[0] == item_id:
                        self.balance_cy_input.setValue(m[4] or 0.0)  # opening_balance_cy
                        self.balance_py_input.setValue(m[5] or 0.0)  # opening_balance_py
                        break
                # Set parent major
                for i in range(self.parent_major_combo.count()):
                    if self.parent_major_combo.itemData(i) == minor[1]:
                        self.parent_major_combo.setCurrentIndex(i)
                        break
                self.selected_minor_id = item_id
        
        elif item_type == "grouping":
            grouping = Grouping.get_by_id(item_id)
            if grouping:
                # grouping tuple: (grouping_id, minor_head_id, grouping_name, code, description)
                self.name_input.setText(grouping[2])
                self.code_input.setText(grouping[3] or "")
                self.description_input.setText(grouping[4] or "")
                # Get full grouping for balances
                grouping_full = Grouping.get_all(company_id=self.current_company_id)
                for g in grouping_full:
                    if g[0] == item_id:
                        self.balance_cy_input.setValue(g[4] or 0.0)  # opening_balance_cy
                        self.balance_py_input.setValue(g[5] or 0.0)  # opening_balance_py
                        break
                # Set parent minor
                minor = MinorHead.get_by_id(grouping[1])
                if minor:
                    # Set major first
                    for i in range(self.parent_major_combo.count()):
                        if self.parent_major_combo.itemData(i) == minor[1]:
                            self.parent_major_combo.setCurrentIndex(i)
                            break
                    # Then set minor
                    for i in range(self.parent_minor_combo.count()):
                        if self.parent_minor_combo.itemData(i) == grouping[1]:
                            self.parent_minor_combo.setCurrentIndex(i)
                            break
                self.selected_grouping_id = item_id
        
        # Enable update and delete buttons
        self.update_btn.setEnabled(True)
        self.delete_btn.setEnabled(True)
    
    def validate_form(self):
        """Validate form inputs"""
        if not self.current_company_id:
            QMessageBox.warning(self, "Validation Error", "No company selected! Please select a company first.")
            return False
        
        if not self.name_input.text().strip():
            QMessageBox.warning(self, "Validation Error", "Name is required!")
            return False
        
        if self.current_level == "minor":
            if not self.parent_major_combo.currentData():
                QMessageBox.warning(self, "Validation Error", "Please select a Major Head!")
                return False
        
        if self.current_level == "grouping":
            if not self.parent_major_combo.currentData():
                QMessageBox.warning(self, "Validation Error", "Please select a Major Head!")
                return False
            if not self.parent_minor_combo.currentData():
                QMessageBox.warning(self, "Validation Error", "Please select a Minor Head!")
                return False
        
        return True
    
    def add_item(self):
        """Add new item"""
        if not self.validate_form():
            return
        
        name = self.name_input.text().strip()
        code = self.code_input.text().strip()
        description = self.description_input.toPlainText().strip()
        balance_cy = self.balance_cy_input.value()
        balance_py = self.balance_py_input.value()
        
        try:
            if self.current_level == "major":
                MajorHead.create(
                    company_id=self.current_company_id,
                    major_head_name=name,
                    category=code,
                    opening_balance_cy=balance_cy,
                    opening_balance_py=balance_py,
                    description=description
                )
                QMessageBox.information(self, "Success", f"Major Head '{name}' added successfully!")
            
            elif self.current_level == "minor":
                major_id = self.parent_major_combo.currentData()
                MinorHead.create(
                    company_id=self.current_company_id,
                    major_head_id=major_id,
                    minor_head_name=name,
                    opening_balance_cy=balance_cy,
                    opening_balance_py=balance_py,
                    code=code,
                    description=description
                )
                QMessageBox.information(self, "Success", f"Minor Head '{name}' added successfully!")
            
            elif self.current_level == "grouping":
                minor_id = self.parent_minor_combo.currentData()
                Grouping.create(
                    company_id=self.current_company_id,
                    minor_head_id=minor_id,
                    grouping_name=name,
                    opening_balance_cy=balance_cy,
                    opening_balance_py=balance_py,
                    code=code,
                    description=description
                )
                QMessageBox.information(self, "Success", f"Grouping '{name}' added successfully!")
            
            self.clear_form()
            self.load_data()
        
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to add item: {str(e)}")
    
    def update_item(self):
        """Update selected item"""
        if not self.validate_form():
            return
        
        name = self.name_input.text().strip()
        code = self.code_input.text().strip()
        description = self.description_input.toPlainText().strip()
        balance_cy = self.balance_cy_input.value()
        balance_py = self.balance_py_input.value()
        
        try:
            if self.current_level == "major" and self.selected_item_id:
                MajorHead.update(
                    major_head_id=self.selected_item_id,
                    major_head_name=name,
                    category=code,
                    opening_balance_cy=balance_cy,
                    opening_balance_py=balance_py,
                    description=description
                )
                QMessageBox.information(self, "Success", f"Major Head updated successfully!")
            
            elif self.current_level == "minor" and self.selected_item_id:
                major_id = self.parent_major_combo.currentData()
                MinorHead.update(
                    minor_head_id=self.selected_item_id,
                    major_head_id=major_id,
                    minor_head_name=name,
                    opening_balance_cy=balance_cy,
                    opening_balance_py=balance_py,
                    code=code,
                    description=description
                )
                QMessageBox.information(self, "Success", f"Minor Head updated successfully!")
            
            elif self.current_level == "grouping" and self.selected_item_id:
                minor_id = self.parent_minor_combo.currentData()
                Grouping.update(
                    grouping_id=self.selected_item_id,
                    minor_head_id=minor_id,
                    grouping_name=name,
                    opening_balance_cy=balance_cy,
                    opening_balance_py=balance_py,
                    code=code,
                    description=description
                )
                QMessageBox.information(self, "Success", f"Grouping updated successfully!")
            
            self.clear_form()
            self.load_data()
        
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to update item: {str(e)}")
    
    def delete_item(self):
        """Delete selected item"""
        selected_item = self.tree.currentItem()
        if not selected_item:
            QMessageBox.warning(self, "Warning", "Please select an item to delete!")
            return
        
        data = selected_item.data(0, Qt.UserRole)
        if not data:
            return
        
        item_type = data["type"]
        item_id = data["id"]
        item_name = selected_item.text(0)
        
        # Confirm deletion
        reply = QMessageBox.question(
            self, "Confirm Deletion",
            f"Are you sure you want to delete '{item_name}'?\n\n"
            f"This will also delete all child items!",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.No:
            return
        
        try:
            if item_type == "major":
                MajorHead.delete(item_id)
            elif item_type == "minor":
                MinorHead.delete(item_id)
            elif item_type == "grouping":
                Grouping.delete(item_id)
            
            QMessageBox.information(self, "Success", f"'{item_name}' deleted successfully!")
            self.clear_form()
            self.load_data()
        
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to delete item: {str(e)}")
    
    def clear_form(self):
        """Clear form inputs"""
        self.name_input.clear()
        self.code_input.clear()
        self.description_input.clear()
        self.balance_cy_input.setValue(0.0)
        self.balance_py_input.setValue(0.0)
        self.update_btn.setEnabled(False)
        self.delete_btn.setEnabled(False)
        self.selected_major_id = None
        self.selected_minor_id = None
        self.selected_grouping_id = None
        self.selected_item_id = None
    
    def export_to_excel(self):
        """Export master data to Excel"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export Master Data", 
            f"master_data_{os.getpid()}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            wb = openpyxl.Workbook()
            
            # Sheet 1: Major Heads
            ws_major = wb.active
            ws_major.title = "Major Heads"
            ws_major.append(["ID", "Name", "Code", "Description"])
            
            for row in MajorHead.get_all():
                ws_major.append(row)
            
            # Sheet 2: Minor Heads
            ws_minor = wb.create_sheet("Minor Heads")
            ws_minor.append(["ID", "Major Head ID", "Name", "Code", "Description"])
            
            for major in MajorHead.get_all():
                for row in MinorHead.get_all(major[0]):
                    ws_minor.append(row)
            
            # Sheet 3: Groupings
            ws_group = wb.create_sheet("Groupings")
            ws_group.append(["ID", "Minor Head ID", "Name", "Code", "Description"])
            
            for major in MajorHead.get_all():
                for minor in MinorHead.get_all(major[0]):
                    for row in Grouping.get_all(minor[0]):
                        ws_group.append(row)
            
            # Format headers
            for ws in [ws_major, ws_minor, ws_group]:
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            
            wb.save(file_path)
            QMessageBox.information(self, "Success", f"Data exported to:\n{file_path}")
        
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Export failed: {str(e)}")
    
    def import_from_excel(self):
        """Import master data from Excel"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Import Master Data", "",
            "Excel Files (*.xlsx *.xls)"
        )
        
        if not file_path:
            return
        
        try:
            wb = openpyxl.load_workbook(file_path)
            
            imported = {"major": 0, "minor": 0, "grouping": 0}
            
            # Import Major Heads
            if "Major Heads" in wb.sheetnames:
                ws = wb["Major Heads"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[1]:  # Name exists
                        try:
                            MajorHead.create(row[1], row[2] or "", row[3] or "")
                            imported["major"] += 1
                        except:
                            pass  # Skip duplicates
            
            # Import Minor Heads
            if "Minor Heads" in wb.sheetnames:
                ws = wb["Minor Heads"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[1] and row[2]:  # Major ID and Name exist
                        try:
                            MinorHead.create(row[1], row[2], row[3] or "", row[4] or "")
                            imported["minor"] += 1
                        except:
                            pass
            
            # Import Groupings
            if "Groupings" in wb.sheetnames:
                ws = wb["Groupings"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[1] and row[2]:  # Minor ID and Name exist
                        try:
                            Grouping.create(row[1], row[2], row[3] or "", row[4] or "")
                            imported["grouping"] += 1
                        except:
                            pass
            
            self.load_data()
            QMessageBox.information(
                self, "Import Complete",
                f"Successfully imported:\n"
                f"• {imported['major']} Major Heads\n"
                f"• {imported['minor']} Minor Heads\n"
                f"• {imported['grouping']} Groupings"
            )
        
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Import failed: {str(e)}")
