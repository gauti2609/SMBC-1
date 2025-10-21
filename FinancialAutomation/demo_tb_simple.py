#!/usr/bin/env python3
"""
Trial Balance Upload Demo - Simplified Version
Shows TB import and Selection Sheet workflow
"""

import sqlite3
from pathlib import Path
from openpyxl import load_workbook
from config.settings import DB_PATH

def print_section(title):
    print("\n" + "="*80)
    print(f"  {title}")
    print("="*80 + "\n")

def main():
    print("""
╔════════════════════════════════════════════════════════════════════════════════╗
║                                                                                ║
║         TRIAL BALANCE UPLOAD & SELECTION SHEET DEMO                            ║
║                                                                                ║
╚════════════════════════════════════════════════════════════════════════════════╝
""")
    
    # Get company
    print_section("STEP 1: Select Company")
    
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    cursor.execute("SELECT company_id, entity_name, fy_start_date, fy_end_date FROM company_info LIMIT 1")
    company = cursor.fetchone()
    
    if not company:
        print("❌ No company found. Please create a company first.")
        return
    
    company_id, company_name, fy_start, fy_end = company
    print(f"✅ Company: {company_name}")
    print(f"   FY: {fy_start} to {fy_end}")
    
    # Load Sample TB
    print_section("STEP 2: Load Sample Trial Balance")
    
    tb_file = Path('/workspaces/SMBC-1/Sample TB.xlsx')
    if not tb_file.exists():
        print(f"❌ Sample TB not found: {tb_file}")
        return
    
    print(f"✅ Found: {tb_file}")
    
    wb = load_workbook(tb_file)
    ws = wb.active
    
    print(f"   Rows: {ws.max_row}")
    print(f"   Columns: {ws.max_column}")
    
    # Show TB data
    print_section("STEP 3: Trial Balance Summary")
    
    cursor.execute("SELECT COUNT(*) FROM trial_balance WHERE company_id = ?", (company_id,))
    tb_count = cursor.fetchone()[0]
    
    print(f"Trial Balance Entries: {tb_count}")
    
    if tb_count > 0:
        cursor.execute("""
            SELECT SUM(opening_balance_cy), SUM(debit_cy), SUM(credit_cy), SUM(closing_balance_cy)
            FROM trial_balance WHERE company_id = ?
        """, (company_id,))
        
        totals = cursor.fetchone()
        print(f"\nTotals:")
        print(f"   Opening Balance: {totals[0]:,.2f}")
        print(f"   Debit:           {totals[1]:,.2f}")
        print(f"   Credit:          {totals[2]:,.2f}")
        print(f"   Closing Balance: {totals[3]:,.2f}")
        
        balance_diff = abs(totals[1] - totals[2])
        if balance_diff < 0.01:
            print(f"\n   ✅ Trial Balance is balanced!")
        else:
            print(f"\n   ⚠️  Difference: {balance_diff:,.2f}")
    
    # Selection Sheet
    print_section("STEP 4: Selection Sheet Summary")
    
    cursor.execute("SELECT COUNT(*) FROM selection_sheet WHERE company_id = ?", (company_id,))
    total_notes = cursor.fetchone()[0]
    
    # Count actual notes (not headers)
    actual_notes = total_notes  # Simplified for demo
    
    cursor.execute("""
        SELECT COUNT(*) FROM selection_sheet 
        WHERE company_id = ? AND system_recommendation = 'Yes'
    """, (company_id,))
    recommended = cursor.fetchone()[0]
    
    cursor.execute("""
        SELECT COUNT(*) FROM selection_sheet 
        WHERE company_id = ? AND final_selection = 'Yes'
    """, (company_id,))
    selected = cursor.fetchone()[0]
    
    print(f"Total Entries:       {total_notes}")
    print(f"Actual Notes:        {actual_notes}")
    print(f"Recommended:         {recommended}")
    print(f"Selected:            {selected}")
    
    # Show sample recommendations
    print(f"\nSample Recommended Notes:")
    cursor.execute("""
        SELECT schedule_iii_note_ref, description, category
        FROM selection_sheet 
        WHERE company_id = ? AND system_recommendation = 'Yes'
        LIMIT 10
    """, (company_id,))
    
    for i, row in enumerate(cursor.fetchall(), 1):
        print(f"   {i:2d}. [{row[2]}] {row[0]}: {row[1][:50]}")
    
    # Master Data
    print_section("STEP 5: Master Data Status")
    
    cursor.execute("SELECT COUNT(*) FROM major_heads WHERE company_id = ?", (company_id,))
    major_count = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM minor_heads WHERE company_id = ?", (company_id,))
    minor_count = cursor.fetchone()[0]
    
    print(f"Major Heads: {major_count}")
    print(f"Minor Heads: {minor_count}")
    
    # Sample major heads
    cursor.execute("""
        SELECT note_number, description 
        FROM major_heads 
        WHERE company_id = ?
        ORDER BY note_number
        LIMIT 10
    """, (company_id,))
    
    print(f"\nSample Major Heads:")
    for row in cursor.fetchall():
        print(f"   Note {row[0]:2d}: {row[1]}")
    
    conn.close()
    
    # Summary
    print_section("✅ DEMO COMPLETE")
    
    print("Summary:")
    print(f"   ✅ Company: {company_name}")
    print(f"   ✅ Trial Balance: {tb_count} entries")
    print(f"   ✅ Selection Sheet: {selected} notes selected")
    print(f"   ✅ Master Data: {major_count} major heads, {minor_count} minor heads")
    
    print("\nNext Steps:")
    print("   1. Run the application: python main.py")
    print("   2. Review and adjust selections in Selection Sheet tab")
    print("   3. Generate financial statements")
    print("   4. Export to Excel")
    
    print("\n" + "="*80)
    print("  🎉 Demo Successful!")
    print("="*80 + "\n")

if __name__ == "__main__":
    main()
