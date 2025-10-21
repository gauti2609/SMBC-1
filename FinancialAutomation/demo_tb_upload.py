#!/usr/bin/env python3
"""
Trial Balance Upload and Selection Sheet Demo
Demonstrates the complete workflow from TB import to Selection Sheet
"""

import sys
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime

from models.company_info import CompanyInfo
from models.trial_balance import TrialBalance
from models.master_data import MajorHead, MinorHead
from models.selection_sheet import SelectionSheet

def print_header(title):
    """Print formatted header"""
    print("\n" + "="*80)
    print(f"  {title}")
    print("="*80 + "\n")

def print_box(text):
    """Print text in a box"""
    lines = text.split('\n')
    max_len = max(len(line) for line in lines)
    print("\n┌" + "─" * (max_len + 2) + "┐")
    for line in lines:
        print(f"│ {line:<{max_len}} │")
    print("└" + "─" * (max_len + 2) + "┘\n")

def demo_trial_balance_upload_and_selection():
    """Complete demo of TB upload and Selection Sheet"""
    
    print("""
╔════════════════════════════════════════════════════════════════════════════════╗
║                                                                                ║
║         TRIAL BALANCE UPLOAD & SELECTION SHEET DEMO                            ║
║                                                                                ║
╚════════════════════════════════════════════════════════════════════════════════╝
""")
    
    # Step 1: Get Company
    print_header("STEP 1: Select Company")
    
    company = CompanyInfo.get_by_name('Demo Company Ltd')
    if not company:
        print("❌ Demo company not found. Please run demo_db_setup.py first.")
        sys.exit(1)
    
    print(f"✅ Company selected: {company.entity_name}")
    print(f"   Company ID: {company.id}")
    print(f"   FY: {company.fy_start_date} to {company.fy_end_date}")
    
    # Step 2: Load Sample TB
    print_header("STEP 2: Load Sample Trial Balance File")
    
    tb_file = Path('/workspaces/SMBC-1/Sample TB.xlsx')
    if not tb_file.exists():
        print(f"❌ Sample TB file not found: {tb_file}")
        sys.exit(1)
    
    print(f"✅ Found Trial Balance file: {tb_file}")
    
    # Read Excel file
    wb = load_workbook(tb_file)
    ws = wb.active
    
    print(f"\n📊 File Analysis:")
    print(f"   Sheet Name: {ws.title}")
    print(f"   Total Rows: {ws.max_row}")
    print(f"   Total Columns: {ws.max_column}")
    
    # Get headers
    headers = [cell.value for cell in ws[1]]
    print(f"\n   Column Headers:")
    for i, header in enumerate(headers, 1):
        if header:
            print(f"      {i}. {header}")
    
    # Step 3: Import Trial Balance
    print_header("STEP 3: Import Trial Balance Data")
    
    # Clear existing TB data for demo company
    existing = TrialBalance.get_by_company(company.id)
    if existing:
        print(f"⚠️  Found {len(existing)} existing TB entries for this company")
        print("   Clearing old data for fresh demo...")
        # In production, you'd have a delete method
    
    imported_count = 0
    skipped_count = 0
    
    print("\n📥 Importing entries...")
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        ledger_name = row[0]  # Column A
        opening_cy = row[1] if row[1] else 0  # Column B
        debit_cy = row[2] if row[2] else 0  # Column C
        credit_cy = row[3] if row[3] else 0  # Column D
        closing_cy = row[4] if row[4] else 0  # Column E
        closing_py = row[5] if row[5] else 0  # Column F
        
        if not ledger_name or ledger_name.strip() == '':
            skipped_count += 1
            continue
        
        # Auto-assign major head based on keywords
        major_head_id = auto_assign_major_head(company.id, ledger_name)
        
        try:
            TrialBalance.create(
                company_id=company.id,
                ledger_name=ledger_name,
                opening_balance_cy=float(opening_cy),
                debit_cy=float(debit_cy),
                credit_cy=float(credit_cy),
                closing_balance_cy=float(closing_cy),
                closing_balance_py=float(closing_py),
                major_head_id=major_head_id,
                minor_head_id=None,
                notes=None
            )
            imported_count += 1
            
            # Show progress every 50 entries
            if imported_count % 50 == 0:
                print(f"   • Imported {imported_count} entries...")
                
        except Exception as e:
            print(f"   ⚠️  Error importing row {row_idx}: {e}")
            skipped_count += 1
    
    print(f"\n✅ Import complete!")
    print(f"   Successfully imported: {imported_count} entries")
    print(f"   Skipped: {skipped_count} entries")
    
    # Step 4: Show Import Summary
    print_header("STEP 4: Trial Balance Summary")
    
    tb_entries = TrialBalance.get_by_company(company.id)
    
    # Calculate totals
    total_opening = sum(e.opening_balance_cy for e in tb_entries)
    total_debit = sum(e.debit_cy for e in tb_entries)
    total_credit = sum(e.credit_cy for e in tb_entries)
    total_closing = sum(e.closing_balance_cy for e in tb_entries)
    
    print(f"Total Entries: {len(tb_entries)}")
    print(f"\nTotals (in {company.unit_of_measurement}):")
    print(f"   Opening Balance CY:  {total_opening:15,.2f}")
    print(f"   Debit CY:            {total_debit:15,.2f}")
    print(f"   Credit CY:           {total_credit:15,.2f}")
    print(f"   Closing Balance CY:  {total_closing:15,.2f}")
    
    # Check balance
    balance_check = abs(total_debit - total_credit)
    if balance_check < 0.01:
        print(f"\n   ✅ Trial Balance is balanced! (Difference: {balance_check:.2f})")
    else:
        print(f"\n   ⚠️  Trial Balance difference: {balance_check:,.2f}")
    
    # Show mapping status
    mapped_count = sum(1 for e in tb_entries if e.major_head_id is not None)
    unmapped_count = len(tb_entries) - mapped_count
    
    print(f"\nMapping Status:")
    print(f"   Mapped to Major Heads:   {mapped_count:3d} ({mapped_count*100/len(tb_entries):.1f}%)")
    print(f"   Unmapped:                {unmapped_count:3d} ({unmapped_count*100/len(tb_entries):.1f}%)")
    
    # Show sample entries
    print(f"\nSample Trial Balance Entries:")
    for i, entry in enumerate(tb_entries[:5], 1):
        major_head = MajorHead.get_by_id(entry.major_head_id) if entry.major_head_id else None
        major_head_desc = major_head.description if major_head else "Unmapped"
        print(f"   {i}. {entry.ledger_name[:40]:<40} → {major_head_desc}")
    
    # Step 5: Initialize Selection Sheet
    print_header("STEP 5: Initialize Selection Sheet")
    
    # Check if already initialized
    existing_notes = SelectionSheet.get_all_for_company(company.id)
    if existing_notes:
        print(f"⚠️  Selection Sheet already has {len(existing_notes)} entries")
        print("   Using existing entries...")
    else:
        # Initialize with 68 default notes
        SelectionSheet.initialize_default_notes(company.id)
        print("✅ Initialized Selection Sheet with 68 default notes")
    
    # Step 6: Update System Recommendations
    print_header("STEP 6: Analyze Trial Balance & Recommend Notes")
    
    print("🔍 Analyzing Trial Balance data...")
    print("   • Checking account balances")
    print("   • Identifying applicable notes")
    print("   • Generating recommendations")
    
    recommended_count = SelectionSheet.update_system_recommendations(company.id)
    
    print(f"\n✅ Analysis complete!")
    print(f"   System recommended: {recommended_count} notes")
    
    # Step 7: Show Selection Sheet
    print_header("STEP 7: Selection Sheet Summary")
    
    all_notes = SelectionSheet.get_all_for_company(company.id)
    
    # Separate section headers and actual notes
    section_headers = [n for n in all_notes if n.is_section_header]
    actual_notes = [n for n in all_notes if not n.is_section_header]
    
    print(f"Total Entries: {len(all_notes)}")
    print(f"   Section Headers: {len(section_headers)}")
    print(f"   Actual Notes: {len(actual_notes)}")
    
    # Count by category
    categories = {}
    for note in actual_notes:
        cat = note.category
        if cat not in categories:
            categories[cat] = {'total': 0, 'recommended': 0}
        categories[cat]['total'] += 1
        if note.system_recommendation == 'Yes':
            categories[cat]['recommended'] += 1
    
    print(f"\nNotes by Category:")
    for cat in sorted(categories.keys()):
        info = categories[cat]
        print(f"   {cat}: {info['total']} notes ({info['recommended']} recommended)")
    
    # Show recommended notes
    recommended_notes = [n for n in actual_notes if n.system_recommendation == 'Yes']
    
    print(f"\n📋 System Recommended Notes ({len(recommended_notes)}):")
    for i, note in enumerate(recommended_notes[:10], 1):
        print(f"   {i:2d}. [{note.category}] {note.schedule_iii_note_ref}: {note.description[:50]}")
    
    if len(recommended_notes) > 10:
        print(f"   ... and {len(recommended_notes) - 10} more")
    
    # Step 8: Simulate User Selection
    print_header("STEP 8: Simulate User Selections")
    
    print("💡 In the UI, users would:")
    print("   1. Review system recommendations")
    print("   2. Override with Yes/No/Blank dropdowns")
    print("   3. Click 'Update Auto-Numbering'")
    print("   4. Selected notes get sequential numbers (1, 2, 3...)")
    
    # Select first 3 recommended notes as example
    if len(recommended_notes) >= 3:
        print(f"\n📝 Example: Selecting 3 notes manually...")
        
        selections = {
            recommended_notes[0].id: 'Yes',
            recommended_notes[1].id: 'Yes',
            recommended_notes[2].id: 'Yes'
        }
        
        SelectionSheet.bulk_update_user_selections(company.id, selections)
        SelectionSheet.update_auto_numbering(company.id)
        
        print("   ✅ User selections saved")
        print("   ✅ Auto-numbering updated")
    
    # Step 9: Show Final Selection
    print_header("STEP 9: Final Selected Notes")
    
    selected_notes = SelectionSheet.get_selected_notes(company.id)
    
    print(f"✅ Total Selected Notes: {len(selected_notes)}")
    
    if selected_notes:
        print(f"\nFinal Selection (Auto-Numbered):")
        for note in selected_notes[:15]:
            print(f"   Note {note.auto_number:2d}: [{note.category}] {note.schedule_iii_note_ref} - {note.description[:45]}")
        
        if len(selected_notes) > 15:
            print(f"   ... and {len(selected_notes) - 15} more")
    
    # Step 10: Completion Summary
    print_header("✅ DEMO COMPLETE")
    
    print_box(f"""Demo Summary:
• Company: {company.entity_name}
• Trial Balance: {len(tb_entries)} entries imported
• Mapped: {mapped_count}/{len(tb_entries)} entries
• Selection Sheet: {len(all_notes)} total notes
• Recommended: {len(recommended_notes)} notes
• Selected: {len(selected_notes)} notes
""")
    
    print("Next Steps in the Application:")
    print("   1. Review and adjust mappings in Trial Balance tab")
    print("   2. Fine-tune selections in Selection Sheet tab")
    print("   3. Enter additional data (PPE, CWIP, Investments)")
    print("   4. Generate financial statements")
    print("   5. Export to Excel (30-sheet workbook with formulas)")
    
    print("\n" + "="*80)
    print("  🎉 Trial Balance & Selection Sheet Demo Complete!")
    print("="*80 + "\n")

def auto_assign_major_head(company_id, ledger_name):
    """Auto-assign major head based on keywords"""
    ledger_lower = ledger_name.lower()
    
    # Get all major heads for the company
    major_heads = MajorHead.get_all_by_company(company_id)
    
    # Mapping keywords to note numbers
    mappings = {
        'computer|laptop|mobile|monitor|furniture|vehicle': 1,  # PPE
        'receivable|debtor|outstanding': 10,  # Trade Receivables
        'cash|bank': 13,  # Cash
        'advance|prepaid': 6,  # Other Current Assets
        'payable|creditor|vendor': 24,  # Trade Payables
        'loan|borrowing': 20,  # Borrowings
        'capital|equity|reserve': 3,  # Equity
        'revenue|sales|income': None,  # P&L items (not in BS)
        'expense|salary|rent|depreciation': None,  # P&L items
        'deposit|security': 7,  # Other Non-Current Assets
        'investment': 5,  # Investments
        'inventory|stock': 9,  # Inventories
    }
    
    for pattern, note_num in mappings.items():
        if note_num is None:
            continue
        
        keywords = pattern.split('|')
        if any(kw in ledger_lower for kw in keywords):
            # Find major head with this note number
            for mh in major_heads:
                if mh.note_number == note_num:
                    return mh.id
    
    return None

if __name__ == "__main__":
    try:
        demo_trial_balance_upload_and_selection()
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
