"""
Investments Input Form - Schedule III Notes 3, 4, 13, 14
Non-Current Investments (Notes 3, 4) and Current Investments (Notes 13, 14)
"""
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QLabel, QMessageBox, QHeaderView, QDialog, QFormLayout,
    QLineEdit, QComboBox, QCheckBox, QSpinBox, QDoubleSpinBox, QTabWidget
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor, QFont
from models.investments import Investment
from datetime import datetime
import traceback


class InvestmentDialog(QDialog):
    """Dialog to add new investment"""
    
    def __init__(self, classification, parent=None):
        super().__init__(parent)
        self.classification = classification
        self.setWindowTitle(f"Add {classification} Investment")
        self.setModal(True)
        self.resize(450, 300)
        
        layout = QFormLayout(self)
        
        # Investment particulars
        self.particulars_input = QLineEdit()
        self.particulars_input.setPlaceholderText("e.g., ABC Ltd - Equity Shares")
        layout.addRow("Investment Name:", self.particulars_input)
        
        # Investment type
        self.type_combo = QComboBox()
        self.type_combo.addItems([
            Investment.TYPE_SUBSIDIARY,
            Investment.TYPE_ASSOCIATE,
            Investment.TYPE_JOINT_VENTURE,
            Investment.TYPE_EQUITY,
            Investment.TYPE_PREFERENCE,
            Investment.TYPE_DEBT,
            Investment.TYPE_MUTUAL_FUND,
            Investment.TYPE_GOVERNMENT_SECURITIES,
            Investment.TYPE_OTHERS
        ])
        layout.addRow("Investment Type:", self.type_combo)
        
        # Is quoted
        self.quoted_check = QCheckBox("Listed/Quoted in Stock Exchange")
        layout.addRow("", self.quoted_check)
        
        # Buttons
        btn_layout = QHBoxLayout()
        
        add_btn = QPushButton("Add")
        add_btn.clicked.connect(self.accept)
        
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        
        btn_layout.addWidget(add_btn)
        btn_layout.addWidget(cancel_btn)
        
        layout.addRow(btn_layout)
    
    def get_investment_data(self):
        """Return entered investment data"""
        return {
            'particulars': self.particulars_input.text().strip(),
            'investment_type': self.type_combo.currentText(),
            'is_quoted': self.quoted_check.isChecked()
        }


class InvestmentTableWidget(QWidget):
    """Table widget for investments with specific classification"""
    
    # Column indices
    COL_PARTICULARS = 0
    COL_TYPE = 1
    COL_QUOTED = 2
    COL_QTY_CY = 3
    COL_QTY_PY = 4
    COL_COST_CY = 5
    COL_COST_PY = 6
    COL_FAIR_VALUE_CY = 7
    COL_FAIR_VALUE_PY = 8
    COL_CARRYING_CY = 9
    COL_CARRYING_PY = 10
    COL_MARKET_CY = 11
    COL_MARKET_PY = 12
    COL_INV_ID = 13  # Hidden
    COL_MODIFIED = 14  # Hidden
    
    def __init__(self, classification, parent=None):
        super().__init__(parent)
        self.classification = classification
        self.company_id = None
        self.parent_form = parent
        self.init_ui()
    
    def init_ui(self):
        """Initialize the UI"""
        layout = QVBoxLayout(self)
        
        # Action buttons
        btn_layout = QHBoxLayout()
        
        self.add_btn = QPushButton("➕ Add Investment")
        self.add_btn.clicked.connect(self.add_investment)
        
        self.delete_btn = QPushButton("🗑️ Delete Selected")
        self.delete_btn.clicked.connect(self.delete_selected)
        
        btn_layout.addWidget(self.add_btn)
        btn_layout.addWidget(self.delete_btn)
        btn_layout.addStretch()
        
        layout.addLayout(btn_layout)
        
        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(15)  # 15 columns (13 visible + 2 hidden)
        
        headers = [
            "Investment Particulars",
            "Type",
            "Quoted",
            "Qty\n(CY)",
            "Qty\n(PY)",
            "Cost\n(CY)",
            "Cost\n(PY)",
            "Fair Value\n(CY)",
            "Fair Value\n(PY)",
            "Carrying Amt\n(CY)",
            "Carrying Amt\n(PY)",
            "Market Value\n(CY)",
            "Market Value\n(PY)",
            "Inv ID",
            "Modified"
        ]
        self.table.setHorizontalHeaderLabels(headers)
        
        # Hide ID and Modified columns
        self.table.setColumnHidden(self.COL_INV_ID, True)
        self.table.setColumnHidden(self.COL_MODIFIED, True)
        
        # Column widths
        self.table.setColumnWidth(self.COL_PARTICULARS, 250)
        self.table.setColumnWidth(self.COL_TYPE, 150)
        self.table.setColumnWidth(self.COL_QUOTED, 80)
        for col in range(self.COL_QTY_CY, self.COL_MARKET_PY + 1):
            self.table.setColumnWidth(col, 120)
        
        # Stretch last visible column
        self.table.horizontalHeader().setStretchLastSection(True)
        
        # Enable sorting
        self.table.setSortingEnabled(True)
        
        # Connect cell change
        self.table.itemChanged.connect(self.on_item_changed)
        
        layout.addWidget(self.table)
        
        # Summary section
        summary_layout = QHBoxLayout()
        
        self.summary_label = QLabel(f"Total {self.classification} Investments:")
        self.summary_label.setStyleSheet("font-weight: bold; font-size: 11pt;")
        
        self.total_cy_label = QLabel("CY: ₹ 0.00")
        self.total_cy_label.setStyleSheet("color: #2196F3; font-size: 11pt; font-weight: bold;")
        
        self.total_py_label = QLabel("PY: ₹ 0.00")
        self.total_py_label.setStyleSheet("color: #FF9800; font-size: 11pt; font-weight: bold;")
        
        summary_layout.addWidget(self.summary_label)
        summary_layout.addWidget(self.total_cy_label)
        summary_layout.addWidget(self.total_py_label)
        summary_layout.addStretch()
        
        layout.addLayout(summary_layout)
    
    def set_company(self, company_id: int):
        """Set the company and load investment data"""
        self.company_id = company_id
        self.load_data()
    
    def load_data(self):
        """Load investment data for the company"""
        if not self.company_id:
            return
        
        try:
            # Block signals
            self.table.blockSignals(True)
            
            # Clear table
            self.table.setRowCount(0)
            
            # Load investments
            investments = Investment.get_all_by_company(self.company_id, self.classification)
            
            for inv in investments:
                self.add_investment_row(inv)
            
            # Update summary
            self.update_summary()
            
            # Unblock signals
            self.table.blockSignals(False)
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load investments:\n{str(e)}")
    
    def add_investment_row(self, inv: Investment):
        """Add an investment row to the table"""
        row = self.table.rowCount()
        self.table.insertRow(row)
        
        # Particulars (editable)
        particulars_item = QTableWidgetItem(inv.investment_particulars)
        self.table.setItem(row, self.COL_PARTICULARS, particulars_item)
        
        # Type (editable via dropdown would be better, but text for now)
        type_item = QTableWidgetItem(inv.investment_type)
        self.table.setItem(row, self.COL_TYPE, type_item)
        
        # Quoted
        quoted_item = QTableWidgetItem("Yes" if inv.is_quoted else "No")
        quoted_item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
        self.table.setItem(row, self.COL_QUOTED, quoted_item)
        
        # Quantities (editable)
        self.set_int_item(row, self.COL_QTY_CY, inv.quantity_cy)
        self.set_int_item(row, self.COL_QTY_PY, inv.quantity_py)
        
        # Amounts (editable)
        self.set_amount_item(row, self.COL_COST_CY, inv.cost_cy)
        self.set_amount_item(row, self.COL_COST_PY, inv.cost_py)
        self.set_amount_item(row, self.COL_FAIR_VALUE_CY, inv.fair_value_cy)
        self.set_amount_item(row, self.COL_FAIR_VALUE_PY, inv.fair_value_py)
        self.set_amount_item(row, self.COL_CARRYING_CY, inv.carrying_amount_cy)
        self.set_amount_item(row, self.COL_CARRYING_PY, inv.carrying_amount_py)
        self.set_amount_item(row, self.COL_MARKET_CY, inv.market_value_cy)
        self.set_amount_item(row, self.COL_MARKET_PY, inv.market_value_py)
        
        # Hidden columns
        inv_id_item = QTableWidgetItem(str(inv.investment_id or ""))
        self.table.setItem(row, self.COL_INV_ID, inv_id_item)
        
        modified_item = QTableWidgetItem("0")
        self.table.setItem(row, self.COL_MODIFIED, modified_item)
    
    def set_amount_item(self, row: int, col: int, value: float):
        """Set an editable amount cell"""
        item = QTableWidgetItem(f"{value:,.2f}")
        item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.table.setItem(row, col, item)
    
    def set_int_item(self, row: int, col: int, value: int):
        """Set an editable integer cell"""
        item = QTableWidgetItem(str(value))
        item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.table.setItem(row, col, item)
    
    def on_item_changed(self, item):
        """Handle cell changes"""
        # Mark as modified
        modified_item = self.table.item(item.row(), self.COL_MODIFIED)
        if modified_item:
            modified_item.setText("1")
        
        # Update summary if carrying amount changed
        if item.column() in [self.COL_CARRYING_CY, self.COL_CARRYING_PY]:
            self.update_summary()
    
    def get_float_value(self, row: int, col: int) -> float:
        """Get float value from a cell"""
        item = self.table.item(row, col)
        if item:
            try:
                return float(item.text().replace(',', ''))
            except ValueError:
                return 0.0
        return 0.0
    
    def get_int_value(self, row: int, col: int) -> int:
        """Get int value from a cell"""
        item = self.table.item(row, col)
        if item:
            try:
                return int(item.text().replace(',', ''))
            except ValueError:
                return 0
        return 0
    
    def update_summary(self):
        """Update the summary totals"""
        total_cy = 0.0
        total_py = 0.0
        
        for row in range(self.table.rowCount()):
            total_cy += self.get_float_value(row, self.COL_CARRYING_CY)
            total_py += self.get_float_value(row, self.COL_CARRYING_PY)
        
        self.total_cy_label.setText(f"CY: ₹ {total_cy:,.2f}")
        self.total_py_label.setText(f"PY: ₹ {total_py:,.2f}")
    
    def add_investment(self):
        """Add a new investment"""
        dialog = InvestmentDialog(self.classification, self)
        if dialog.exec_() == QDialog.Accepted:
            inv_data = dialog.get_investment_data()
            
            if not inv_data['particulars']:
                QMessageBox.warning(self, "Warning", "Please enter investment name.")
                return
            
            # Create empty investment object
            inv = Investment(
                investment_particulars=inv_data['particulars'],
                classification=self.classification,
                investment_type=inv_data['investment_type'],
                is_quoted=inv_data['is_quoted']
            )
            
            # Add to table
            self.add_investment_row(inv)
            
            # Mark as new
            row = self.table.rowCount() - 1
            modified_item = self.table.item(row, self.COL_MODIFIED)
            if modified_item:
                modified_item.setText("1")
    
    def delete_selected(self):
        """Delete selected investments"""
        selected_rows = set(item.row() for item in self.table.selectedItems())
        
        if not selected_rows:
            QMessageBox.warning(self, "Warning", "Please select investments to delete.")
            return
        
        reply = QMessageBox.question(
            self,
            "Confirm Delete",
            f"Delete {len(selected_rows)} investment(s)?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            for row in sorted(selected_rows, reverse=True):
                # Get investment ID
                inv_id_item = self.table.item(row, self.COL_INV_ID)
                if inv_id_item and inv_id_item.text():
                    inv_id = int(inv_id_item.text())
                    Investment.delete(inv_id)
                
                # Remove row
                self.table.removeRow(row)
            
            self.update_summary()
            QMessageBox.information(self, "Success", "Investments deleted successfully!")
    
    def save_all(self):
        """Save all investments"""
        if not self.company_id:
            return 0
        
        saved_count = 0
        
        try:
            for row in range(self.table.rowCount()):
                # Check if modified
                modified_item = self.table.item(row, self.COL_MODIFIED)
                if not modified_item or modified_item.text() != "1":
                    continue
                
                # Get data
                particulars = self.table.item(row, self.COL_PARTICULARS).text()
                inv_type = self.table.item(row, self.COL_TYPE).text()
                quoted_item = self.table.item(row, self.COL_QUOTED)
                is_quoted = quoted_item.text().lower() == "yes" if quoted_item else False
                
                qty_cy = self.get_int_value(row, self.COL_QTY_CY)
                qty_py = self.get_int_value(row, self.COL_QTY_PY)
                
                cost_cy = self.get_float_value(row, self.COL_COST_CY)
                cost_py = self.get_float_value(row, self.COL_COST_PY)
                fair_cy = self.get_float_value(row, self.COL_FAIR_VALUE_CY)
                fair_py = self.get_float_value(row, self.COL_FAIR_VALUE_PY)
                carrying_cy = self.get_float_value(row, self.COL_CARRYING_CY)
                carrying_py = self.get_float_value(row, self.COL_CARRYING_PY)
                market_cy = self.get_float_value(row, self.COL_MARKET_CY)
                market_py = self.get_float_value(row, self.COL_MARKET_PY)
                
                # Get investment ID
                inv_id_item = self.table.item(row, self.COL_INV_ID)
                inv_id = int(inv_id_item.text()) if inv_id_item and inv_id_item.text() else None
                
                if inv_id:
                    # Update existing
                    Investment.update(
                        investment_id=inv_id,
                        investment_particulars=particulars,
                        classification=self.classification,
                        investment_type=inv_type,
                        is_quoted=is_quoted,
                        quantity_cy=qty_cy,
                        quantity_py=qty_py,
                        cost_cy=cost_cy,
                        cost_py=cost_py,
                        fair_value_cy=fair_cy,
                        fair_value_py=fair_py,
                        carrying_amount_cy=carrying_cy,
                        carrying_amount_py=carrying_py,
                        market_value_cy=market_cy,
                        market_value_py=market_py
                    )
                else:
                    # Create new
                    new_id = Investment.create(
                        company_id=self.company_id,
                        investment_particulars=particulars,
                        classification=self.classification,
                        investment_type=inv_type,
                        is_quoted=is_quoted,
                        quantity_cy=qty_cy,
                        quantity_py=qty_py,
                        cost_cy=cost_cy,
                        cost_py=cost_py,
                        fair_value_cy=fair_cy,
                        fair_value_py=fair_py,
                        carrying_amount_cy=carrying_cy,
                        carrying_amount_py=carrying_py,
                        market_value_cy=market_cy,
                        market_value_py=market_py
                    )
                    # Update investment ID in table
                    inv_id_item.setText(str(new_id))
                
                # Reset modified flag
                modified_item.setText("0")
                saved_count += 1
                
        except Exception as e:
            raise e
        
        return saved_count


class InvestmentsInputForm(QWidget):
    """Investments Input Form with Non-Current and Current tabs"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.company_id = None
        self.init_ui()
    
    def init_ui(self):
        """Initialize the UI"""
        layout = QVBoxLayout(self)
        
        # Title
        title = QLabel("💰 Investments - Schedule III Notes 3, 4, 13, 14")
        title_font = QFont()
        title_font.setPointSize(14)
        title_font.setBold(True)
        title.setFont(title_font)
        layout.addWidget(title)
        
        # Info text
        info = QLabel("Track investments with Current Year (CY) and Previous Year (PY) comparative data")
        info.setStyleSheet("color: #666; margin-bottom: 10px;")
        layout.addWidget(info)
        
        # Action buttons
        btn_layout = QHBoxLayout()
        
        self.save_btn = QPushButton("💾 Save All")
        self.save_btn.clicked.connect(self.save_all)
        self.save_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold;")
        
        self.export_btn = QPushButton("📤 Export Schedule III")
        self.export_btn.clicked.connect(self.export_schedule_iii)
        
        btn_layout.addStretch()
        btn_layout.addWidget(self.export_btn)
        btn_layout.addWidget(self.save_btn)
        
        layout.addLayout(btn_layout)
        
        # Create tabs for Non-Current and Current
        self.tabs = QTabWidget()
        
        # Non-Current Investments (Notes 3, 4)
        self.non_current_tab = InvestmentTableWidget(Investment.NON_CURRENT, self)
        self.tabs.addTab(self.non_current_tab, "📊 Non-Current Investments (Notes 3, 4)")
        
        # Current Investments (Notes 13, 14)
        self.current_tab = InvestmentTableWidget(Investment.CURRENT, self)
        self.tabs.addTab(self.current_tab, "💵 Current Investments (Notes 13, 14)")
        
        layout.addWidget(self.tabs)
    
    def set_company(self, company_id: int):
        """Set the company and load data"""
        self.company_id = company_id
        self.non_current_tab.set_company(company_id)
        self.current_tab.set_company(company_id)
    
    def save_all(self):
        """Save all investments from both tabs"""
        if not self.company_id:
            QMessageBox.warning(self, "Warning", "No company selected.")
            return
        
        try:
            # Save both tabs
            non_current_saved = self.non_current_tab.save_all()
            current_saved = self.current_tab.save_all()
            
            total_saved = non_current_saved + current_saved
            
            if total_saved > 0:
                QMessageBox.information(self, "Success", f"Saved {total_saved} investment(s) successfully!")
            else:
                QMessageBox.information(self, "Info", "No changes to save.")
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save:\n{str(e)}\n{traceback.format_exc()}")
    
    def export_schedule_iii(self):
        """Export investments in Schedule III format"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from PyQt5.QtWidgets import QFileDialog
            
            # Get file path
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Export Schedule III",
                f"Investments_Schedule_III_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
            
            # Create workbook
            wb = Workbook()
            
            # Non-Current sheet
            ws_nc = wb.active
            ws_nc.title = "Non-Current Investments"
            self.export_tab_to_sheet(ws_nc, self.non_current_tab, "NON-CURRENT INVESTMENTS - NOTES 3, 4")
            
            # Current sheet
            ws_c = wb.create_sheet("Current Investments")
            self.export_tab_to_sheet(ws_c, self.current_tab, "CURRENT INVESTMENTS - NOTES 13, 14")
            
            # Save
            wb.save(file_path)
            
            QMessageBox.information(self, "Success", f"Exported to:\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Export failed:\n{str(e)}")
    
    def export_tab_to_sheet(self, ws, tab_widget, title):
        """Export a tab to a worksheet"""
        # Title
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:M1')
        
        # Headers
        headers = [
            "Investment Particulars",
            "Type",
            "Quoted",
            "Quantity (CY)",
            "Quantity (PY)",
            "Cost (CY)",
            "Cost (PY)",
            "Fair Value (CY)",
            "Fair Value (PY)",
            "Carrying Amount (CY)",
            "Carrying Amount (PY)",
            "Market Value (CY)",
            "Market Value (PY)"
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        table = tab_widget.table
        for row in range(table.rowCount()):
            for col in range(13):  # 13 visible columns
                item = table.item(row, col)
                if item:
                    ws.cell(row=row+4, column=col+1, value=item.text())
