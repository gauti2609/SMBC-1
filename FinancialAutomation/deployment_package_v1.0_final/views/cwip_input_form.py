"""
Capital Work in Progress (CWIP) Input Form - Schedule III Note 2
"""
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QLabel, QMessageBox, QHeaderView, QDialog, QFormLayout,
    QLineEdit, QDateEdit
)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QColor, QFont
from models.cwip import CWIP
from datetime import datetime
import traceback


class ProjectDialog(QDialog):
    """Dialog to add new CWIP project"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Add CWIP Project")
        self.setModal(True)
        self.resize(400, 150)
        
        layout = QFormLayout(self)
        
        # Project name input
        self.project_input = QLineEdit()
        self.project_input.setPlaceholderText("e.g., Factory Building Extension")
        layout.addRow("Project Name:", self.project_input)
        
        # Start date
        self.start_date = QDateEdit()
        self.start_date.setCalendarPopup(True)
        self.start_date.setDate(QDate.currentDate())
        layout.addRow("Start Date:", self.start_date)
        
        # Expected completion date
        self.completion_date = QDateEdit()
        self.completion_date.setCalendarPopup(True)
        self.completion_date.setDate(QDate.currentDate().addYears(1))
        layout.addRow("Expected Completion:", self.completion_date)
        
        # Buttons
        btn_layout = QHBoxLayout()
        
        add_btn = QPushButton("Add")
        add_btn.clicked.connect(self.accept)
        
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        
        btn_layout.addWidget(add_btn)
        btn_layout.addWidget(cancel_btn)
        
        layout.addRow(btn_layout)
    
    def get_project_data(self):
        """Return entered project data"""
        return {
            'project_name': self.project_input.text().strip(),
            'project_start_date': self.start_date.date().toString("yyyy-MM-dd"),
            'expected_completion_date': self.completion_date.date().toString("yyyy-MM-dd")
        }


class CWIPInputForm(QWidget):
    """CWIP Input Form with CY & PY comparative data"""
    
    # Column indices
    COL_PROJECT = 0
    COL_OPENING_CY = 1
    COL_ADDITIONS_CY = 2
    COL_CAPITALIZED_CY = 3
    COL_CLOSING_CY = 4
    COL_OPENING_PY = 5
    COL_ADDITIONS_PY = 6
    COL_CAPITALIZED_PY = 7
    COL_CLOSING_PY = 8
    COL_START_DATE = 9
    COL_COMPLETION_DATE = 10
    COL_CWIP_ID = 11  # Hidden
    COL_MODIFIED = 12  # Hidden
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.company_id = None
        self.init_ui()
    
    def init_ui(self):
        """Initialize the UI"""
        layout = QVBoxLayout(self)
        
        # Title
        title = QLabel("🚧 Capital Work in Progress (CWIP) - Schedule III Note 2")
        title_font = QFont()
        title_font.setPointSize(14)
        title_font.setBold(True)
        title.setFont(title_font)
        layout.addWidget(title)
        
        # Info text
        info = QLabel("Track ongoing construction projects with Current Year (CY) and Previous Year (PY) comparative data")
        info.setStyleSheet("color: #666; margin-bottom: 10px;")
        layout.addWidget(info)
        
        # Action buttons
        btn_layout = QHBoxLayout()
        
        self.add_btn = QPushButton("➕ Add Project")
        self.add_btn.clicked.connect(self.add_project)
        
        self.delete_btn = QPushButton("🗑️ Delete Selected")
        self.delete_btn.clicked.connect(self.delete_selected)
        
        self.save_btn = QPushButton("💾 Save All")
        self.save_btn.clicked.connect(self.save_all)
        self.save_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold;")
        
        self.import_btn = QPushButton("📥 Import Excel")
        self.import_btn.clicked.connect(self.import_excel)
        
        self.export_btn = QPushButton("📤 Export Schedule III")
        self.export_btn.clicked.connect(self.export_schedule_iii)
        
        btn_layout.addWidget(self.add_btn)
        btn_layout.addWidget(self.delete_btn)
        btn_layout.addStretch()
        btn_layout.addWidget(self.import_btn)
        btn_layout.addWidget(self.export_btn)
        btn_layout.addWidget(self.save_btn)
        
        layout.addLayout(btn_layout)
        
        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(13)  # 13 columns (11 visible + 2 hidden)
        
        headers = [
            "Project Name",
            "Opening Balance\n(CY)",
            "Additions\n(CY)",
            "Capitalized\n(CY)",
            "Closing Balance\n(CY)",
            "Opening Balance\n(PY)",
            "Additions\n(PY)",
            "Capitalized\n(PY)",
            "Closing Balance\n(PY)",
            "Start Date",
            "Expected\nCompletion",
            "CWIP ID",
            "Modified"
        ]
        self.table.setHorizontalHeaderLabels(headers)
        
        # Hide ID and Modified columns
        self.table.setColumnHidden(self.COL_CWIP_ID, True)
        self.table.setColumnHidden(self.COL_MODIFIED, True)
        
        # Column widths
        self.table.setColumnWidth(self.COL_PROJECT, 200)
        for col in range(self.COL_OPENING_CY, self.COL_COMPLETION_DATE):
            self.table.setColumnWidth(col, 120)
        
        # Stretch last visible column
        self.table.horizontalHeader().setStretchLastSection(True)
        
        # Enable sorting
        self.table.setSortingEnabled(True)
        
        # Connect cell change to auto-calculate
        self.table.itemChanged.connect(self.on_item_changed)
        
        layout.addWidget(self.table)
        
        # Summary section
        summary_layout = QHBoxLayout()
        
        self.summary_label = QLabel("Total CWIP:")
        self.summary_label.setStyleSheet("font-weight: bold; font-size: 12pt;")
        
        self.total_cy_label = QLabel("CY: ₹ 0.00")
        self.total_cy_label.setStyleSheet("color: #2196F3; font-size: 12pt; font-weight: bold;")
        
        self.total_py_label = QLabel("PY: ₹ 0.00")
        self.total_py_label.setStyleSheet("color: #FF9800; font-size: 12pt; font-weight: bold;")
        
        summary_layout.addWidget(self.summary_label)
        summary_layout.addWidget(self.total_cy_label)
        summary_layout.addWidget(self.total_py_label)
        summary_layout.addStretch()
        
        layout.addLayout(summary_layout)
    
    def set_company(self, company_id: int):
        """Set the company and load CWIP data"""
        self.company_id = company_id
        self.load_data()
    
    def load_data(self):
        """Load CWIP data for the company"""
        if not self.company_id:
            return
        
        try:
            # Block signals to prevent triggering calculations during load
            self.table.blockSignals(True)
            
            # Clear table
            self.table.setRowCount(0)
            
            # Load CWIP data
            cwip_list = CWIP.get_all_by_company(self.company_id)
            
            for cwip in cwip_list:
                self.add_cwip_row(cwip)
            
            # Update summary
            self.update_summary()
            
            # Unblock signals
            self.table.blockSignals(False)
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load CWIP data:\n{str(e)}")
    
    def add_cwip_row(self, cwip: CWIP):
        """Add a CWIP row to the table"""
        row = self.table.rowCount()
        self.table.insertRow(row)
        
        # Project name (editable)
        project_item = QTableWidgetItem(cwip.project_name)
        self.table.setItem(row, self.COL_PROJECT, project_item)
        
        # Current Year amounts (editable)
        self.set_amount_item(row, self.COL_OPENING_CY, cwip.opening_balance_cy)
        self.set_amount_item(row, self.COL_ADDITIONS_CY, cwip.additions_cy)
        self.set_amount_item(row, self.COL_CAPITALIZED_CY, cwip.capitalized_cy)
        
        # CY Closing (calculated, read-only)
        closing_cy = cwip.calculate_closing_balance_cy()
        self.set_calculated_item(row, self.COL_CLOSING_CY, closing_cy)
        
        # Previous Year amounts (editable)
        self.set_amount_item(row, self.COL_OPENING_PY, cwip.opening_balance_py)
        self.set_amount_item(row, self.COL_ADDITIONS_PY, cwip.additions_py)
        self.set_amount_item(row, self.COL_CAPITALIZED_PY, cwip.capitalized_py)
        
        # PY Closing (calculated, read-only)
        closing_py = cwip.calculate_closing_balance_py()
        self.set_calculated_item(row, self.COL_CLOSING_PY, closing_py)
        
        # Dates (editable)
        start_date_item = QTableWidgetItem(cwip.project_start_date or "")
        self.table.setItem(row, self.COL_START_DATE, start_date_item)
        
        completion_date_item = QTableWidgetItem(cwip.expected_completion_date or "")
        self.table.setItem(row, self.COL_COMPLETION_DATE, completion_date_item)
        
        # Hidden columns
        cwip_id_item = QTableWidgetItem(str(cwip.cwip_id or ""))
        self.table.setItem(row, self.COL_CWIP_ID, cwip_id_item)
        
        modified_item = QTableWidgetItem("0")
        self.table.setItem(row, self.COL_MODIFIED, modified_item)
    
    def set_amount_item(self, row: int, col: int, value: float):
        """Set an editable amount cell"""
        item = QTableWidgetItem(f"{value:,.2f}")
        item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.table.setItem(row, col, item)
    
    def set_calculated_item(self, row: int, col: int, value: float):
        """Set a calculated (read-only) cell"""
        item = QTableWidgetItem(f"{value:,.2f}")
        item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Make read-only
        item.setBackground(QColor("#F0F0F0"))  # Gray background
        item.setForeground(QColor("#2196F3"))  # Blue text
        self.table.setItem(row, col, item)
    
    def on_item_changed(self, item):
        """Handle cell changes and recalculate"""
        if item.column() in [
            self.COL_OPENING_CY, self.COL_ADDITIONS_CY, self.COL_CAPITALIZED_CY,
            self.COL_OPENING_PY, self.COL_ADDITIONS_PY, self.COL_CAPITALIZED_PY
        ]:
            # Mark as modified
            modified_item = self.table.item(item.row(), self.COL_MODIFIED)
            if modified_item:
                modified_item.setText("1")
            
            # Recalculate row
            self.recalculate_row(item.row())
            
            # Update summary
            self.update_summary()
    
    def recalculate_row(self, row: int):
        """Recalculate closing balances for a row"""
        try:
            # Block signals to prevent recursive calls
            self.table.blockSignals(True)
            
            # Get CY amounts
            opening_cy = self.get_float_value(row, self.COL_OPENING_CY)
            additions_cy = self.get_float_value(row, self.COL_ADDITIONS_CY)
            capitalized_cy = self.get_float_value(row, self.COL_CAPITALIZED_CY)
            
            # Calculate CY closing
            closing_cy = opening_cy + additions_cy - capitalized_cy
            self.set_calculated_item(row, self.COL_CLOSING_CY, closing_cy)
            
            # Get PY amounts
            opening_py = self.get_float_value(row, self.COL_OPENING_PY)
            additions_py = self.get_float_value(row, self.COL_ADDITIONS_PY)
            capitalized_py = self.get_float_value(row, self.COL_CAPITALIZED_PY)
            
            # Calculate PY closing
            closing_py = opening_py + additions_py - capitalized_py
            self.set_calculated_item(row, self.COL_CLOSING_PY, closing_py)
            
            # Unblock signals
            self.table.blockSignals(False)
            
        except Exception as e:
            self.table.blockSignals(False)
            print(f"Error recalculating row: {e}")
    
    def get_float_value(self, row: int, col: int) -> float:
        """Get float value from a cell"""
        item = self.table.item(row, col)
        if item:
            try:
                # Remove commas and convert
                return float(item.text().replace(',', ''))
            except ValueError:
                return 0.0
        return 0.0
    
    def update_summary(self):
        """Update the summary totals"""
        total_cy = 0.0
        total_py = 0.0
        
        for row in range(self.table.rowCount()):
            total_cy += self.get_float_value(row, self.COL_CLOSING_CY)
            total_py += self.get_float_value(row, self.COL_CLOSING_PY)
        
        self.total_cy_label.setText(f"CY: ₹ {total_cy:,.2f}")
        self.total_py_label.setText(f"PY: ₹ {total_py:,.2f}")
    
    def add_project(self):
        """Add a new CWIP project"""
        dialog = ProjectDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            project_data = dialog.get_project_data()
            
            if not project_data['project_name']:
                QMessageBox.warning(self, "Warning", "Please enter a project name.")
                return
            
            # Create empty CWIP object
            cwip = CWIP(
                project_name=project_data['project_name'],
                project_start_date=project_data['project_start_date'],
                expected_completion_date=project_data['expected_completion_date']
            )
            
            # Add to table
            self.add_cwip_row(cwip)
            
            # Mark as new (cwip_id will be empty)
            row = self.table.rowCount() - 1
            modified_item = self.table.item(row, self.COL_MODIFIED)
            if modified_item:
                modified_item.setText("1")
    
    def delete_selected(self):
        """Delete selected CWIP projects"""
        selected_rows = set(item.row() for item in self.table.selectedItems())
        
        if not selected_rows:
            QMessageBox.warning(self, "Warning", "Please select projects to delete.")
            return
        
        reply = QMessageBox.question(
            self,
            "Confirm Delete",
            f"Delete {len(selected_rows)} project(s)?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            for row in sorted(selected_rows, reverse=True):
                # Get CWIP ID
                cwip_id_item = self.table.item(row, self.COL_CWIP_ID)
                if cwip_id_item and cwip_id_item.text():
                    cwip_id = int(cwip_id_item.text())
                    CWIP.delete(cwip_id)
                
                # Remove row
                self.table.removeRow(row)
            
            self.update_summary()
            QMessageBox.information(self, "Success", "Projects deleted successfully!")
    
    def save_all(self):
        """Save all CWIP projects"""
        if not self.company_id:
            QMessageBox.warning(self, "Warning", "No company selected.")
            return
        
        try:
            saved_count = 0
            
            for row in range(self.table.rowCount()):
                # Check if modified
                modified_item = self.table.item(row, self.COL_MODIFIED)
                if not modified_item or modified_item.text() != "1":
                    continue
                
                # Get data
                project_name = self.table.item(row, self.COL_PROJECT).text()
                
                opening_cy = self.get_float_value(row, self.COL_OPENING_CY)
                additions_cy = self.get_float_value(row, self.COL_ADDITIONS_CY)
                capitalized_cy = self.get_float_value(row, self.COL_CAPITALIZED_CY)
                
                opening_py = self.get_float_value(row, self.COL_OPENING_PY)
                additions_py = self.get_float_value(row, self.COL_ADDITIONS_PY)
                capitalized_py = self.get_float_value(row, self.COL_CAPITALIZED_PY)
                
                start_date_item = self.table.item(row, self.COL_START_DATE)
                start_date = start_date_item.text() if start_date_item else None
                
                completion_date_item = self.table.item(row, self.COL_COMPLETION_DATE)
                completion_date = completion_date_item.text() if completion_date_item else None
                
                # Get CWIP ID
                cwip_id_item = self.table.item(row, self.COL_CWIP_ID)
                cwip_id = int(cwip_id_item.text()) if cwip_id_item and cwip_id_item.text() else None
                
                if cwip_id:
                    # Update existing
                    CWIP.update(
                        cwip_id=cwip_id,
                        project_name=project_name,
                        opening_balance_cy=opening_cy,
                        additions_cy=additions_cy,
                        capitalized_cy=capitalized_cy,
                        opening_balance_py=opening_py,
                        additions_py=additions_py,
                        capitalized_py=capitalized_py,
                        project_start_date=start_date,
                        expected_completion_date=completion_date
                    )
                else:
                    # Create new
                    new_id = CWIP.create(
                        company_id=self.company_id,
                        project_name=project_name,
                        opening_balance_cy=opening_cy,
                        additions_cy=additions_cy,
                        capitalized_cy=capitalized_cy,
                        opening_balance_py=opening_py,
                        additions_py=additions_py,
                        capitalized_py=capitalized_py,
                        project_start_date=start_date,
                        expected_completion_date=completion_date
                    )
                    # Update CWIP ID in table
                    cwip_id_item.setText(str(new_id))
                
                # Reset modified flag
                modified_item.setText("0")
                saved_count += 1
            
            if saved_count > 0:
                QMessageBox.information(self, "Success", f"Saved {saved_count} project(s) successfully!")
            else:
                QMessageBox.information(self, "Info", "No changes to save.")
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save:\n{str(e)}\n{traceback.format_exc()}")
    
    def import_excel(self):
        """Import CWIP data from Excel"""
        QMessageBox.information(self, "Coming Soon", "Excel import functionality will be implemented soon.")
    
    def export_schedule_iii(self):
        """Export CWIP in Schedule III format"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from PyQt5.QtWidgets import QFileDialog
            
            # Get file path
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Export Schedule III",
                f"CWIP_Schedule_III_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "CWIP - Note 2"
            
            # Title
            ws['A1'] = "CAPITAL WORK IN PROGRESS - SCHEDULE III NOTE 2"
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:K1')
            
            # Headers
            headers = [
                "Project Name",
                "Opening Balance (CY)",
                "Additions (CY)",
                "Capitalized (CY)",
                "Closing Balance (CY)",
                "Opening Balance (PY)",
                "Additions (PY)",
                "Capitalized (PY)",
                "Closing Balance (PY)",
                "Start Date",
                "Expected Completion"
            ]
            
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Data
            for row in range(self.table.rowCount()):
                for col in range(11):  # 11 visible columns
                    item = self.table.item(row, col)
                    if item:
                        ws.cell(row=row+4, column=col+1, value=item.text())
            
            # Save
            wb.save(file_path)
            
            QMessageBox.information(self, "Success", f"Exported to:\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Export failed:\n{str(e)}")
