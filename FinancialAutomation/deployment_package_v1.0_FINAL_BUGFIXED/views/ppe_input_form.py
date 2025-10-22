"""PPE Input Form - Property, Plant & Equipment (Schedule III Note 1)"""

from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
                             QLabel, QTableWidget, QTableWidgetItem, QMessageBox,
                             QHeaderView, QGroupBox, QComboBox, QDoubleSpinBox,
                             QSpinBox, QLineEdit, QDialog, QFormLayout, QFileDialog,
                             QAbstractItemView)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont, QColor
from models.ppe import PPE
import openpyxl
from openpyxl.styles import Font as ExcelFont, PatternFill, Alignment
import os

class PPEInputForm(QWidget):
    """Property, Plant & Equipment input form with CY & PY support"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_window = parent
        self.init_ui()
        self.load_data()
    
    @property
    def current_company_id(self):
        """Get current company ID from parent window"""
        if hasattr(self.parent_window, 'current_company_id'):
            return self.parent_window.current_company_id
        return None
    
    def init_ui(self):
        """Initialize UI"""
        layout = QVBoxLayout()
        
        # Header
        header = QLabel("🏭 Property, Plant & Equipment (Schedule III Note 1)")
        header.setFont(QFont("Bookman Old Style", 14, QFont.Bold))
        header.setStyleSheet("color: #2c3e50; padding: 10px;")
        layout.addWidget(header)
        
        # Info
        info = QLabel(
            "Enter asset-wise movements for Current Year (CY) and Previous Year (PY). "
            "Closing values are auto-calculated. Data feeds directly into Schedule III Balance Sheet."
        )
        info.setWordWrap(True)
        info.setStyleSheet("padding: 5px; background-color: #e8f5e9; border-radius: 5px;")
        layout.addWidget(info)
        
        # Action buttons
        button_layout = self.create_action_buttons()
        layout.addLayout(button_layout)
        
        # PPE Table
        self.create_ppe_table()
        layout.addWidget(self.table)
        
        # Summary section
        summary_group = self.create_summary_section()
        layout.addWidget(summary_group)
        
        # Bottom action buttons
        bottom_layout = QHBoxLayout()
        
        save_btn = QPushButton("💾 Save All Changes")
        save_btn.clicked.connect(self.save_all)
        save_btn.setStyleSheet("padding: 10px 20px; background-color: #4CAF50; color: white; font-weight: bold;")
        bottom_layout.addWidget(save_btn)
        
        refresh_btn = QPushButton("🔄 Refresh")
        refresh_btn.clicked.connect(self.load_data)
        refresh_btn.setStyleSheet("padding: 10px 20px;")
        bottom_layout.addWidget(refresh_btn)
        
        bottom_layout.addStretch()
        layout.addLayout(bottom_layout)
        
        self.setLayout(layout)
    
    def create_action_buttons(self):
        """Create action buttons"""
        layout = QHBoxLayout()
        
        add_btn = QPushButton("➕ Add Asset Class")
        add_btn.clicked.connect(self.add_asset_class)
        add_btn.setStyleSheet("padding: 8px 16px; background-color: #2196F3; color: white;")
        layout.addWidget(add_btn)
        
        delete_btn = QPushButton("🗑️ Delete Selected")
        delete_btn.clicked.connect(self.delete_selected)
        delete_btn.setStyleSheet("padding: 8px 16px; background-color: #f44336; color: white;")
        layout.addWidget(delete_btn)
        
        import_btn = QPushButton("📥 Import Excel")
        import_btn.clicked.connect(self.import_excel)
        layout.addWidget(import_btn)
        
        export_btn = QPushButton("📤 Export Schedule III")
        export_btn.clicked.connect(self.export_schedule_iii)
        layout.addWidget(export_btn)
        
        layout.addStretch()
        
        return layout
    
    def create_ppe_table(self):
        """Create PPE data table"""
        self.table = QTableWidget()
        self.table.setColumnCount(23)
        
        headers = [
            "Asset Class",
            # CY - Gross Block
            "Opening GB (CY)", "Additions (CY)", "Disposals (CY)", "Closing GB (CY)",
            # CY - Depreciation
            "Opening Dep (CY)", "Dep for Year (CY)", "Dep on Disposals (CY)", "Closing Dep (CY)",
            # CY - Net Block
            "Net Block (CY)",
            # PY - Gross Block
            "Opening GB (PY)", "Additions (PY)", "Disposals (PY)", "Closing GB (PY)",
            # PY - Depreciation
            "Opening Dep (PY)", "Dep for Year (PY)", "Dep on Disposals (PY)", "Closing Dep (PY)",
            # PY - Net Block
            "Net Block (PY)",
            # Metadata
            "Dep Rate (%)", "Useful Life (Years)", "PPE ID", "Modified"
        ]
        
        self.table.setHorizontalHeaderLabels(headers)
        self.table.verticalHeader().setVisible(False)
        
        # Set column widths
        self.table.setColumnWidth(0, 200)  # Asset Class
        for i in range(1, 19):  # Amount columns
            self.table.setColumnWidth(i, 120)
        self.table.setColumnWidth(19, 80)  # Dep Rate
        self.table.setColumnWidth(20, 120)  # Useful Life
        self.table.setColumnHidden(21, True)  # PPE ID (hidden)
        self.table.setColumnHidden(22, True)  # Modified flag (hidden)
        
        # Style
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("""
            QTableWidget {
                gridline-color: #d0d0d0;
                background-color: white;
            }
            QTableWidget::item:selected {
                background-color: #b3d9ff;
            }
            QHeaderView::section {
                background-color: #2c3e50;
                color: white;
                padding: 8px;
                border: 1px solid #34495e;
                font-weight: bold;
            }
        """)
        
        self.table.itemChanged.connect(self.on_item_changed)
    
    def create_summary_section(self):
        """Create summary section showing totals"""
        group = QGroupBox("📊 Summary - Total PPE")
        layout = QHBoxLayout()
        
        self.summary_label_cy = QLabel("Current Year Net Block: ₹ 0.00")
        self.summary_label_cy.setFont(QFont("Bookman Old Style", 11, QFont.Bold))
        self.summary_label_cy.setStyleSheet("padding: 10px; background-color: #e3f2fd; color: #1565c0;")
        layout.addWidget(self.summary_label_cy)
        
        self.summary_label_py = QLabel("Previous Year Net Block: ₹ 0.00")
        self.summary_label_py.setFont(QFont("Bookman Old Style", 11, QFont.Bold))
        self.summary_label_py.setStyleSheet("padding: 10px; background-color: #fff3e0; color: #e65100;")
        layout.addWidget(self.summary_label_py)
        
        group.setLayout(layout)
        return group
    
    def load_data(self):
        """Load PPE data from database"""
        if not self.current_company_id:
            self.update_status("⚠️ No company selected", error=True)
            return
        
        self.table.blockSignals(True)
        self.table.setRowCount(0)
        
        try:
            ppe_list = PPE.get_all_by_company(self.current_company_id)
            
            for ppe in ppe_list:
                self.add_ppe_row(ppe)
            
            self.calculate_summary()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load PPE data:\n{str(e)}")
        finally:
            self.table.blockSignals(False)
    
    def add_ppe_row(self, ppe=None):
        """Add a row to the table"""
        row = self.table.rowCount()
        self.table.insertRow(row)
        
        if ppe:
            # Asset Class
            self.table.setItem(row, 0, QTableWidgetItem(ppe.asset_class))
            
            # CY - Gross Block
            self.set_numeric_item(row, 1, ppe.opening_gross_block_cy)
            self.set_numeric_item(row, 2, ppe.additions_cy)
            self.set_numeric_item(row, 3, ppe.disposals_gross_cy)
            self.set_numeric_item(row, 4, ppe.calculate_closing_gross_block_cy(), readonly=True, calculated=True)
            
            # CY - Depreciation
            self.set_numeric_item(row, 5, ppe.opening_acc_depreciation_cy)
            self.set_numeric_item(row, 6, ppe.depreciation_for_year_cy)
            self.set_numeric_item(row, 7, ppe.acc_depr_on_disposals_cy)
            self.set_numeric_item(row, 8, ppe.calculate_closing_acc_depreciation_cy(), readonly=True, calculated=True)
            
            # CY - Net Block
            self.set_numeric_item(row, 9, ppe.calculate_net_block_cy(), readonly=True, calculated=True)
            
            # PY - Gross Block
            self.set_numeric_item(row, 10, ppe.opening_gross_block_py)
            self.set_numeric_item(row, 11, ppe.additions_py)
            self.set_numeric_item(row, 12, ppe.disposals_gross_py)
            self.set_numeric_item(row, 13, ppe.calculate_closing_gross_block_py(), readonly=True, calculated=True)
            
            # PY - Depreciation
            self.set_numeric_item(row, 14, ppe.opening_acc_depreciation_py)
            self.set_numeric_item(row, 15, ppe.depreciation_for_year_py)
            self.set_numeric_item(row, 16, ppe.acc_depr_on_disposals_py)
            self.set_numeric_item(row, 17, ppe.calculate_closing_acc_depreciation_py(), readonly=True, calculated=True)
            
            # PY - Net Block
            self.set_numeric_item(row, 18, ppe.calculate_net_block_py(), readonly=True, calculated=True)
            
            # Metadata
            self.set_numeric_item(row, 19, ppe.depreciation_rate)
            self.set_numeric_item(row, 20, ppe.useful_life_years)
            
            # Hidden columns
            self.table.setItem(row, 21, QTableWidgetItem(str(ppe.ppe_id)))  # PPE ID
            self.table.setItem(row, 22, QTableWidgetItem("0"))  # Modified flag
        else:
            # Empty row for new entry
            self.table.setItem(row, 0, QTableWidgetItem(""))
            for col in range(1, 21):
                if col not in [4, 8, 9, 13, 17, 18]:  # Skip calculated columns
                    self.set_numeric_item(row, col, 0.0)
                else:
                    self.set_numeric_item(row, col, 0.0, readonly=True, calculated=True)
            self.table.setItem(row, 21, QTableWidgetItem(""))  # No ID for new row
            self.table.setItem(row, 22, QTableWidgetItem("1"))  # Mark as modified
    
    def set_numeric_item(self, row, col, value, readonly=False, calculated=False):
        """Set a numeric item in the table"""
        item = QTableWidgetItem(f"{value:,.2f}")
        item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        
        if readonly:
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            if calculated:
                item.setBackground(QColor("#f0f0f0"))
                item.setForeground(QColor("#1976d2"))
        
        self.table.setItem(row, col, item)
    
    def on_item_changed(self, item):
        """Handle item change - recalculate closing values"""
        row = item.row()
        col = item.column()
        
        # Skip if it's a calculated column
        if col in [4, 8, 9, 13, 17, 18, 21, 22]:
            return
        
        # Mark row as modified
        modified_item = self.table.item(row, 22)
        if modified_item:
            modified_item.setText("1")
        
        # Recalculate closing values for this row
        self.recalculate_row(row)
        self.calculate_summary()
    
    def recalculate_row(self, row):
        """Recalculate closing values for a row"""
        self.table.blockSignals(True)
        
        try:
            # CY Closing Gross Block = Opening + Additions - Disposals
            opening_gb_cy = self.get_numeric_value(row, 1)
            additions_cy = self.get_numeric_value(row, 2)
            disposals_cy = self.get_numeric_value(row, 3)
            closing_gb_cy = opening_gb_cy + additions_cy - disposals_cy
            self.set_numeric_item(row, 4, closing_gb_cy, readonly=True, calculated=True)
            
            # CY Closing Depreciation = Opening + For Year - On Disposals
            opening_dep_cy = self.get_numeric_value(row, 5)
            dep_year_cy = self.get_numeric_value(row, 6)
            dep_disposals_cy = self.get_numeric_value(row, 7)
            closing_dep_cy = opening_dep_cy + dep_year_cy - dep_disposals_cy
            self.set_numeric_item(row, 8, closing_dep_cy, readonly=True, calculated=True)
            
            # CY Net Block = Closing GB - Closing Dep
            net_block_cy = closing_gb_cy - closing_dep_cy
            self.set_numeric_item(row, 9, net_block_cy, readonly=True, calculated=True)
            
            # PY Closing Gross Block
            opening_gb_py = self.get_numeric_value(row, 10)
            additions_py = self.get_numeric_value(row, 11)
            disposals_py = self.get_numeric_value(row, 12)
            closing_gb_py = opening_gb_py + additions_py - disposals_py
            self.set_numeric_item(row, 13, closing_gb_py, readonly=True, calculated=True)
            
            # PY Closing Depreciation
            opening_dep_py = self.get_numeric_value(row, 14)
            dep_year_py = self.get_numeric_value(row, 15)
            dep_disposals_py = self.get_numeric_value(row, 16)
            closing_dep_py = opening_dep_py + dep_year_py - dep_disposals_py
            self.set_numeric_item(row, 17, closing_dep_py, readonly=True, calculated=True)
            
            # PY Net Block
            net_block_py = closing_gb_py - closing_dep_py
            self.set_numeric_item(row, 18, net_block_py, readonly=True, calculated=True)
            
        except Exception as e:
            print(f"Error recalculating row {row}: {e}")
        finally:
            self.table.blockSignals(False)
    
    def get_numeric_value(self, row, col):
        """Get numeric value from table cell"""
        item = self.table.item(row, col)
        if item and item.text():
            try:
                # Remove commas and convert to float
                return float(item.text().replace(',', ''))
            except:
                return 0.0
        return 0.0
    
    def calculate_summary(self):
        """Calculate and display summary totals"""
        total_net_cy = 0.0
        total_net_py = 0.0
        
        for row in range(self.table.rowCount()):
            total_net_cy += self.get_numeric_value(row, 9)
            total_net_py += self.get_numeric_value(row, 18)
        
        self.summary_label_cy.setText(f"Current Year Net Block: ₹ {total_net_cy:,.2f}")
        self.summary_label_py.setText(f"Previous Year Net Block: ₹ {total_net_py:,.2f}")
    
    def add_asset_class(self):
        """Add a new asset class"""
        if not self.current_company_id:
            QMessageBox.warning(self, "No Company", "Please select a company first")
            return
        
        dialog = AssetClassDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            asset_class = dialog.asset_class_input.text().strip()
            if asset_class:
                self.add_ppe_row()
                row = self.table.rowCount() - 1
                self.table.setItem(row, 0, QTableWidgetItem(asset_class))
                self.table.setItem(row, 22, QTableWidgetItem("1"))  # Mark as modified
    
    def delete_selected(self):
        """Delete selected rows"""
        selected_rows = set(item.row() for item in self.table.selectedItems())
        
        if not selected_rows:
            QMessageBox.warning(self, "No Selection", "Please select rows to delete")
            return
        
        reply = QMessageBox.question(
            self, "Confirm Delete",
            f"Delete {len(selected_rows)} asset class(es)?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.No:
            return
        
        # Delete from database and table
        for row in sorted(selected_rows, reverse=True):
            ppe_id_item = self.table.item(row, 21)
            if ppe_id_item and ppe_id_item.text():
                try:
                    PPE.delete(int(ppe_id_item.text()))
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Failed to delete: {str(e)}")
            
            self.table.removeRow(row)
        
        self.calculate_summary()
        QMessageBox.information(self, "Success", f"Deleted {len(selected_rows)} asset class(es)")
    
    def save_all(self):
        """Save all changes to database"""
        if not self.current_company_id:
            QMessageBox.warning(self, "No Company", "Please select a company first")
            return
        
        saved_count = 0
        error_count = 0
        
        for row in range(self.table.rowCount()):
            modified_item = self.table.item(row, 22)
            if not modified_item or modified_item.text() != "1":
                continue  # Skip unmodified rows
            
            asset_class_item = self.table.item(row, 0)
            if not asset_class_item or not asset_class_item.text().strip():
                continue  # Skip empty rows
            
            try:
                asset_class = asset_class_item.text().strip()
                ppe_id_item = self.table.item(row, 21)
                ppe_id = int(ppe_id_item.text()) if ppe_id_item and ppe_id_item.text() else None
                
                # Get all values
                data = {
                    'asset_class': asset_class,
                    'opening_gross_block_cy': self.get_numeric_value(row, 1),
                    'additions_cy': self.get_numeric_value(row, 2),
                    'disposals_gross_cy': self.get_numeric_value(row, 3),
                    'opening_acc_depreciation_cy': self.get_numeric_value(row, 5),
                    'depreciation_for_year_cy': self.get_numeric_value(row, 6),
                    'acc_depr_on_disposals_cy': self.get_numeric_value(row, 7),
                    'opening_gross_block_py': self.get_numeric_value(row, 10),
                    'additions_py': self.get_numeric_value(row, 11),
                    'disposals_gross_py': self.get_numeric_value(row, 12),
                    'opening_acc_depreciation_py': self.get_numeric_value(row, 14),
                    'depreciation_for_year_py': self.get_numeric_value(row, 15),
                    'acc_depr_on_disposals_py': self.get_numeric_value(row, 16),
                    'depreciation_rate': self.get_numeric_value(row, 19),
                    'useful_life_years': int(self.get_numeric_value(row, 20))
                }
                
                if ppe_id:
                    # Update existing
                    PPE.update(ppe_id, **data)
                else:
                    # Create new
                    new_id = PPE.create(self.current_company_id, **data)
                    self.table.setItem(row, 21, QTableWidgetItem(str(new_id)))
                
                # Mark as saved
                self.table.setItem(row, 22, QTableWidgetItem("0"))
                saved_count += 1
                
            except Exception as e:
                error_count += 1
                print(f"Error saving row {row}: {e}")
        
        if error_count > 0:
            QMessageBox.warning(
                self, "Partial Save",
                f"Saved {saved_count} entries, {error_count} errors occurred"
            )
        else:
            QMessageBox.information(self, "Success", f"Saved {saved_count} PPE entries successfully!")
        
        self.load_data()
    
    def import_excel(self):
        """Import PPE data from Excel"""
        QMessageBox.information(
            self, "Import Excel",
            "Excel import feature will be implemented.\n\n"
            "Format: Asset Class, Opening GB CY, Additions CY, etc."
        )
    
    def export_schedule_iii(self):
        """Export data in Schedule III format"""
        if not self.current_company_id:
            QMessageBox.warning(self, "No Company", "Please select a company first")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export Schedule III - PPE",
            f"Schedule_III_Note_1_PPE_{self.current_company_id}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            schedule_data = PPE.get_schedule_iii_format(self.current_company_id)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "PPE - Note 1"
            
            # Title
            ws.merge_cells('A1:V1')
            ws['A1'] = "Schedule III - Note 1: Property, Plant and Equipment"
            ws['A1'].font = ExcelFont(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Headers
            headers = [
                "Asset Class",
                "Opening", "Additions", "Disposals", "Closing",
                "Opening", "For Year", "On Disposals", "Closing",
                "Opening", "Closing",
                "Opening", "Additions", "Disposals", "Closing",
                "Opening", "For Year", "On Disposals", "Closing",
                "Opening", "Closing"
            ]
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx, value=header)
                cell.font = ExcelFont(bold=True)
                cell.fill = PatternFill(start_color="4472C4", fill_type="solid")
            
            # Data rows
            row_idx = 4
            for data in schedule_data:
                ws.cell(row=row_idx, column=1, value=data['asset_class'])
                # CY Gross Block
                ws.cell(row=row_idx, column=2, value=data['gross_block_opening_cy'])
                ws.cell(row=row_idx, column=3, value=data['gross_block_additions_cy'])
                ws.cell(row=row_idx, column=4, value=data['gross_block_disposals_cy'])
                ws.cell(row=row_idx, column=5, value=data['gross_block_closing_cy'])
                # CY Depreciation
                ws.cell(row=row_idx, column=6, value=data['depreciation_opening_cy'])
                ws.cell(row=row_idx, column=7, value=data['depreciation_for_year_cy'])
                ws.cell(row=row_idx, column=8, value=data['depreciation_on_disposals_cy'])
                ws.cell(row=row_idx, column=9, value=data['depreciation_closing_cy'])
                # CY Net Block
                ws.cell(row=row_idx, column=10, value=data['net_block_opening_cy'])
                ws.cell(row=row_idx, column=11, value=data['net_block_closing_cy'])
                # PY Gross Block
                ws.cell(row=row_idx, column=12, value=data['gross_block_opening_py'])
                ws.cell(row=row_idx, column=13, value=data['gross_block_additions_py'])
                ws.cell(row=row_idx, column=14, value=data['gross_block_disposals_py'])
                ws.cell(row=row_idx, column=15, value=data['gross_block_closing_py'])
                # PY Depreciation
                ws.cell(row=row_idx, column=16, value=data['depreciation_opening_py'])
                ws.cell(row=row_idx, column=17, value=data['depreciation_for_year_py'])
                ws.cell(row=row_idx, column=18, value=data['depreciation_on_disposals_py'])
                ws.cell(row=row_idx, column=19, value=data['depreciation_closing_py'])
                # PY Net Block
                ws.cell(row=row_idx, column=20, value=data['net_block_opening_py'])
                ws.cell(row=row_idx, column=21, value=data['net_block_closing_py'])
                
                row_idx += 1
            
            wb.save(file_path)
            QMessageBox.information(self, "Success", f"Exported to:\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Export failed:\n{str(e)}")
    
    def update_status(self, message, error=False):
        """Update status (placeholder)"""
        print(message)


class AssetClassDialog(QDialog):
    """Dialog for adding new asset class"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Add Asset Class")
        self.setModal(True)
        self.init_ui()
    
    def init_ui(self):
        """Initialize UI"""
        layout = QVBoxLayout()
        
        form_layout = QFormLayout()
        
        self.asset_class_input = QLineEdit()
        self.asset_class_input.setPlaceholderText("e.g., Land, Building, Plant & Machinery")
        form_layout.addRow("Asset Class Name:", self.asset_class_input)
        
        layout.addLayout(form_layout)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        ok_btn = QPushButton("OK")
        ok_btn.clicked.connect(self.accept)
        ok_btn.setDefault(True)
        button_layout.addWidget(ok_btn)
        
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
